# -*- coding: utf-8 -*-
"""
Sistema de Honorários — controle de contratos, parcelas e recebimentos.

Stack: Streamlit + PostgreSQL (Supabase), deploy no Railway.

Organização do arquivo:
    1.  Imports e constantes
    2.  Configuração da página e estilo
    3.  Configuração (variáveis de ambiente / secrets)
    4.  Camada de banco (pool, transações, cache)
    5.  Schema (DDL, colunas novas, índices)
    6.  Keep-alive do Supabase
    7.  Utilitários (datas, validação, formatação)
    8.  Exportação (Excel / PDF)
    9.  Recibos
    10. Autenticação
    11. Páginas
    12. Roteamento
"""

from __future__ import annotations

# =============================================================================
# 1. IMPORTS E CONSTANTES
# =============================================================================
import atexit
import base64
import hmac
import io
import os
import pathlib
import re
import urllib.parse
from contextlib import contextmanager
from datetime import date, datetime, timedelta, timezone
from typing import Any, Callable, Sequence

import altair as alt
import pandas as pd
import psycopg2
import psycopg2.extensions
import psycopg2.pool
import streamlit as st
from psycopg2.extras import RealDictCursor, execute_values

# Dependências opcionais: a aplicação continua de pé sem elas, apenas
# desabilitando o recurso correspondente (evita quebrar o deploy inteiro).
try:
    from fpdf import FPDF

    PDF_DISPONIVEL = True
except Exception:  # pragma: no cover
    FPDF = object  # type: ignore[assignment,misc]
    PDF_DISPONIVEL = False

try:
    from apscheduler.schedulers.background import BackgroundScheduler

    SCHEDULER_DISPONIVEL = True
except Exception:  # pragma: no cover
    BackgroundScheduler = None  # type: ignore[assignment,misc]
    SCHEDULER_DISPONIVEL = False

# Fuso horário do escritório. O Railway roda em UTC — sem isso, recibos e
# cálculos de atraso saem com até 3 horas (ou um dia) de diferença.
try:
    from zoneinfo import ZoneInfo

    FUSO = ZoneInfo("America/Sao_Paulo")
except Exception:  # pragma: no cover - fallback se o tzdata não existir
    FUSO = timezone(timedelta(hours=-3))

APP_TITULO = "Sistema de Honorários"
APP_ICONE = "⚖️"
ESCRITORIO = "Maciel Freitas Advocacia"

MENU = [
    "📊 Dashboard",
    "➕ Novo Contrato",
    "💰 Pagamentos",
    "📂 Meus Contratos",
    "📁 Arquivados",
    "⚙️ Gestão",
]

# Data no padrão brasileiro (dia primeiro), definida num lugar só.
# Para trocar a barra por hífen em todo o sistema, basta alterar estas duas
# linhas de forma coerente: "%d-%m-%Y" e "DD-MM-YYYY".
FORMATO_DATA = "%d/%m/%Y"
FORMATO_DATA_WIDGET = "DD/MM/YYYY"

# Locale do Vega: faz o eixo e as dicas do gráfico saírem em pt-BR
# ("R$ 15.000" em vez de "$15,000").
LOCALE_VEGA = {
    "number": {
        "decimal": ",",
        "thousands": ".",
        "grouping": [3],
        "currency": ["R$ ", ""],
    }
}

FORMAS_PAGAMENTO = ["Pix", "Dinheiro", "Transferência", "Cartão", "Boleto"]
STATUS_TUTELA = ["Pendente", "Deferido", "Indeferido", "Parcial"]
TUTELA_COM_REDUCAO = ("Deferido", "Parcial")

# Consultas ficam em cache por até CACHE_TTL segundos. Qualquer escrita feita
# pelo próprio sistema invalida o cache na hora (ver `_invalidar_cache`).
CACHE_TTL = 120

MAX_TENTATIVAS_LOGIN = 5
BLOQUEIO_LOGIN_SEG = 60

# psycopg2 devolve NUMERIC como Decimal, o que quebraria as contas em float
# espalhadas pelo sistema. Converter para float mantém tudo compatível caso
# um dia as colunas REAL sejam migradas para NUMERIC(14,2) (ver migracoes.sql).
psycopg2.extensions.register_type(
    psycopg2.extensions.new_type(
        psycopg2.extensions.DECIMAL.values,
        "DEC2FLOAT",
        lambda valor, _cur: float(valor) if valor is not None else None,
    )
)


class ErroBanco(RuntimeError):
    """Falha ao falar com o banco — sempre exibida ao usuário."""


class ConfiguracaoAusente(RuntimeError):
    """Variável de ambiente / secret obrigatória não encontrada."""


# =============================================================================
# 2. CONFIGURAÇÃO DA PÁGINA E ESTILO
# =============================================================================
st.set_page_config(
    page_title=APP_TITULO,
    layout="wide",
    page_icon=APP_ICONE,
    initial_sidebar_state="expanded",
)

# Paleta e tipografia do escritório (gmfreitas.com.br). O tema base vem do
# .streamlit/config.toml; aqui vai o que o config não alcança.
st.markdown(
    """
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,500;9..144,600&family=Source+Serif+4:opsz,wght@8..60,400;8..60,600&display=swap">
    <style>
        :root {
            --teal:        #035E70;
            --teal-escuro: #024250;
            --teal-tinta:  #012B34;
            --teal-claro:  #0A7F95;
            --teal-suave:  #E7F0F1;
            --tinta:       #1C2A2E;
            --tinta-suave: #43545A;
            --cinza:       #8B9296;
            --creme:       #F7F5F0;
            --papel:       #FBFAF7;
            --linha:       rgba(3, 94, 112, .16);
            --sombra:      0 2px 14px rgba(2, 43, 52, .06);

            --ok:      #10603F;  --ok-fundo:      #E2F2EA;
            --alerta:  #8A5A00;  --alerta-fundo:  #FBF1D8;
            --critico: #A8201A;  --critico-fundo: #FBE9E7;
        }

        /* Largura contida: a tabela deixa de se esticar de ponta a ponta num
           monitor grande, que era o que mais atrapalhava a leitura. */
        .block-container {
            max-width: 1240px;
            padding-top: 1.4rem;
            padding-bottom: 4rem;
        }

        html, body, [class*="css"] { font-family: "Source Serif 4", Georgia, serif; }
        h1, h2, h3, h4 {
            font-family: "Fraunces", Georgia, serif !important;
            color: var(--teal-tinta);
            letter-spacing: -.01em;
        }
        h1 { font-size: 1.9rem !important; }
        h2 { font-size: 1.42rem !important; }
        h3 { font-size: 1.12rem !important; }

        /* Cabeçalho da marca */
        .marca {
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 20px;
            padding: 4px 0 14px;
            border-bottom: 2px solid var(--teal);
            margin-bottom: 22px;
            flex-wrap: wrap;
        }
        .marca img { height: 38px; width: auto; }
        .marca-titulo {
            font-family: "Fraunces", Georgia, serif;
            font-size: 1.02rem;
            color: var(--teal-escuro);
            text-align: right;
            line-height: 1.35;
        }
        .marca-titulo small {
            display: block;
            font-family: "Source Serif 4", Georgia, serif;
            font-size: .78rem;
            color: var(--cinza);
        }

        /* Indicadores: mais baixos e mais densos que o padrão do Streamlit,
           que gastava um terço da tela em três números. */
        [data-testid="stMetric"] {
            background: #FFFFFF;
            border: 1px solid var(--linha);
            border-radius: 12px;
            padding: 12px 16px;
            box-shadow: var(--sombra);
        }
        [data-testid="stMetricLabel"] p {
            font-size: .76rem !important;
            text-transform: uppercase;
            letter-spacing: .07em;
            color: var(--cinza) !important;
        }
        [data-testid="stMetricValue"] {
            font-family: "Fraunces", Georgia, serif;
            font-size: 1.42rem !important;
            color: var(--teal-tinta);
            font-variant-numeric: tabular-nums;
        }

        button[kind="primary"] { width: 100%; height: 3em; font-weight: 600; letter-spacing: .01em; }
        button[kind="secondary"] { border-color: var(--linha); }

        .caixa {
            background: #FFFFFF;
            border: 1px solid var(--linha);
            border-left: 4px solid var(--realce, var(--teal));
            border-radius: 10px;
            padding: 14px 18px;
            margin-bottom: 14px;
            line-height: 1.65;
            box-shadow: var(--sombra);
        }
        .caixa-cliente { --realce: var(--teal); }
        .caixa-nota    { --realce: #C08A2E; background: #FDFBF5; }
        .caixa-ok      { --realce: var(--ok); }


        section[data-testid="stSidebar"] {
            background: var(--creme);
            border-right: 1px solid var(--linha);
        }
        section[data-testid="stSidebar"] img { max-width: 100%; }

        div[data-testid="stDataFrame"] {
            border: 1px solid var(--linha);
            border-radius: 10px;
            overflow: hidden;
        }

        div[data-testid="stExpander"] {
            border: 1px solid var(--linha);
            border-radius: 10px;
            background: #FFFFFF;
        }

        hr { border-color: var(--linha) !important; }

        /* Some com a barra do Streamlit para o sistema parecer um produto,
           não um script publicado. Esconde só a barra de ferramentas: o
           cabeçalho inteiro não pode sumir porque é nele que fica o botão de
           reabrir a barra lateral — sem ele, quem recolhesse o menu ficaria
           sem como trazê-lo de volta. */
        #MainMenu,
        [data-testid="stToolbar"],
        [data-testid="stDecoration"],
        [data-testid="stStatusWidget"],
        footer { display: none !important; }

        header[data-testid="stHeader"] { background: transparent; height: 0; }
        [data-testid="stSidebarCollapsedControl"] { display: flex !important; }
    </style>
    """,
    unsafe_allow_html=True,
)


# =============================================================================
# 3. CONFIGURAÇÃO (VARIÁVEIS DE AMBIENTE / SECRETS)
# =============================================================================
def _config(env: str, secao: str, chave: str, padrao: str | None = None) -> str:
    """Lê configuração do ambiente (Railway) ou de st.secrets (Streamlit Cloud)."""
    valor = os.environ.get(env)
    if valor:
        return valor
    try:
        valor = st.secrets[secao][chave]
        if valor not in (None, ""):
            return str(valor)
    except Exception:
        pass
    if padrao is not None:
        return padrao
    raise ConfiguracaoAusente(
        f"Configuração ausente: defina a variável de ambiente `{env}` "
        f"ou `[{secao}].{chave}` em .streamlit/secrets.toml."
    )


@st.cache_resource(show_spinner=False)
def _parametros_conexao() -> dict[str, Any]:
    return {
        "host": _config("SUPABASE_HOST", "supabase", "host"),
        "port": int(_config("SUPABASE_PORT", "supabase", "port", "5432")),
        "dbname": _config("SUPABASE_DBNAME", "supabase", "dbname", "postgres"),
        "user": _config("SUPABASE_USER", "supabase", "user"),
        "password": _config("SUPABASE_PASSWORD", "supabase", "password"),
        "sslmode": "require",
        "connect_timeout": 10,
        "keepalives": 1,
        "keepalives_idle": 30,
        "keepalives_interval": 10,
        "keepalives_count": 5,
        "application_name": "sistema-honorarios",
        # Corta consultas travadas antes que elas segurem uma conexão do pool.
        "options": "-c statement_timeout=20000",
    }


# =============================================================================
# 4. CAMADA DE BANCO
# =============================================================================
ERROS_CONEXAO = (psycopg2.OperationalError, psycopg2.InterfaceError)


@st.cache_resource(show_spinner=False)
def _pool() -> psycopg2.pool.ThreadedConnectionPool:
    """Pool compartilhado pelo processo.

    A versão anterior guardava UMA conexão em cache_resource e a compartilhava
    entre todas as sessões: dois usuários simultâneos disputavam a mesma
    transação, e um `rollback()` de um desfazia o trabalho do outro.
    """
    maximo = int(_config("DB_MAX_CONEXOES", "supabase", "max_conexoes", "5"))
    criado = psycopg2.pool.ThreadedConnectionPool(1, max(2, maximo), **_parametros_conexao())
    atexit.register(lambda: _silencioso(criado.closeall))
    return criado


def _silencioso(funcao: Callable[[], Any]) -> None:
    try:
        funcao()
    except Exception:
        pass


def _descartar(pool_, conexao) -> None:
    try:
        pool_.putconn(conexao, close=True)
    except Exception:
        _silencioso(conexao.close)


@contextmanager
def _conexao(verificar: bool = False):
    """Empresta uma conexão do pool e devolve ao final.

    `verificar=True` faz um `SELECT 1` antes de entregar a conexão. Custa um
    round-trip a mais, mas evita retentar uma escrita — o Supabase derruba
    conexões ociosas e retentar um INSERT poderia duplicar o registro.
    """
    # Criar o pool já abre a primeira conexão: se o Supabase estiver fora,
    # o erro nasce aqui e precisa virar ErroBanco para a UI tratá-lo.
    try:
        pool_ = _pool()
    except ConfiguracaoAusente:
        raise
    except Exception as erro:
        raise ErroBanco(f"Não foi possível abrir conexão com o banco: {erro}") from erro

    conexao = None
    ultimo_erro: Exception | None = None

    for _ in range(3 if verificar else 1):
        try:
            candidata = pool_.getconn()
        except Exception as erro:
            ultimo_erro = erro
            continue
        if not verificar:
            conexao = candidata
            break
        try:
            if candidata.closed:
                raise psycopg2.InterfaceError("conexão fechada")
            with candidata.cursor() as cur:
                cur.execute("SELECT 1")
            candidata.rollback()
            conexao = candidata
            break
        except Exception as erro:
            ultimo_erro = erro
            _descartar(pool_, candidata)

    if conexao is None:
        raise ErroBanco(f"Não foi possível obter conexão com o banco: {ultimo_erro}")

    try:
        yield conexao
    finally:
        # putconn faz rollback sozinho se a conexão voltar suja, e fecha a
        # conexão se o socket tiver morrido.
        try:
            pool_.putconn(conexao)
        except Exception:
            _descartar(pool_, conexao)


@contextmanager
def transacao():
    """Agrupa várias operações numa única transação (tudo ou nada).

    Antes, cadastrar um contrato fazia N commits separados: se a criação das
    parcelas falhasse no meio, o contrato ficava salvo sem parcelas.
    """
    with _conexao(verificar=True) as conexao:
        try:
            with conexao.cursor() as cur:
                yield cur
            conexao.commit()
        except Exception:
            _silencioso(conexao.rollback)
            raise
    _invalidar_cache()


def exec_db(query: str, params: Sequence = ()) -> None:
    """Executa uma escrita isolada. Levanta ErroBanco em caso de falha."""
    try:
        with transacao() as cur:
            cur.execute(query, tuple(params))
    except ErroBanco:
        raise
    except Exception as erro:
        raise ErroBanco(str(erro)) from erro


def exec_retorna(query: str, params: Sequence = ()) -> Any:
    """Executa uma escrita com RETURNING e devolve a primeira coluna."""
    try:
        with transacao() as cur:
            cur.execute(query, tuple(params))
            linha = cur.fetchone()
            return linha[0] if linha else None
    except ErroBanco:
        raise
    except Exception as erro:
        raise ErroBanco(str(erro)) from erro


def _select(query: str, params: tuple) -> pd.DataFrame:
    erros: list[Exception] = []
    for tentativa in range(2):
        try:
            with _conexao(verificar=tentativa > 0) as conexao:
                try:
                    with conexao.cursor(cursor_factory=RealDictCursor) as cur:
                        cur.execute(query, params)
                        linhas = cur.fetchall()
                    conexao.commit()
                except Exception:
                    _silencioso(conexao.rollback)
                    raise
            if not linhas:
                return pd.DataFrame()
            df = pd.DataFrame([dict(linha) for linha in linhas])
            # Mantém o comportamento histórico: nulos viram "" para que as
            # checagens de `nulo()` e as formatações continuem valendo.
            return df.where(pd.notnull(df), "")
        except ERROS_CONEXAO as erro:
            # Leitura é idempotente: pode retentar sem risco.
            erros.append(erro)
        except ErroBanco:
            raise
        except Exception as erro:
            raise ErroBanco(str(erro)) from erro
    raise ErroBanco(f"Falha de conexão ao consultar o banco: {erros[-1]}")


@st.cache_resource(show_spinner=False)
def _estado_cache() -> dict[str, int]:
    return {"versao": 0}


def _invalidar_cache() -> None:
    _estado_cache()["versao"] += 1


@st.cache_data(ttl=CACHE_TTL, show_spinner=False, max_entries=128)
def _consulta_cacheada(query: str, params: tuple, versao: int) -> pd.DataFrame:
    return _select(query, params)


def select_db(query: str, params: Sequence = (), cache: bool = True) -> pd.DataFrame:
    """Consulta o banco.

    O Streamlit re-executa o script inteiro a cada clique. Sem cache, uma
    simples troca de cliente no selectbox disparava 4-6 consultas idênticas.
    """
    parametros = tuple(params)
    if not cache:
        return _select(query, parametros)
    return _consulta_cacheada(query, parametros, _estado_cache()["versao"]).copy()


def escalar(query: str, params: Sequence = (), padrao: Any = 0) -> Any:
    df = select_db(query, params)
    if df.empty:
        return padrao
    valor = df.iloc[0, 0]
    return padrao if valor in ("", None) else valor


# =============================================================================
# 5. SCHEMA
# =============================================================================
DDL_CONTRATOS = """
-- Os tipos abaixo espelham o banco de produção, conferido via
-- information_schema. Não é o schema "ideal" — é o real, de propósito: um
-- ambiente de teste criado do zero precisa reproduzir produção, senão não
-- pega os erros que importam. O migracoes.sql cuida de corrigir o resto.
CREATE TABLE IF NOT EXISTS contratos (
    id                      SERIAL PRIMARY KEY,
    cliente                 TEXT NOT NULL,
    cpf_cnpj                TEXT,
    telefone                TEXT,
    valor_total             NUMERIC NOT NULL,
    saldo_devedor           NUMERIC NOT NULL,
    data_contrato           DATE NOT NULL,
    observacoes             TEXT,
    hon_inicial_ativo       TEXT,
    hon_inicial_valor       REAL,
    hon_inicial_parcelado   TEXT,
    hon_inicial_parcelas    INTEGER,
    hon_inicial_vlr_parcela REAL,
    hon_liminar_fixo        REAL,
    hon_liminar_reducao_vlr REAL,
    hon_liminar_reducao_prc INTEGER,
    tutela                  TEXT,
    hon_exito_percentual    REAL,
    hon_exito_fixo          REAL,
    nr_processo             TEXT,
    nr_vara                 TEXT,
    nome_juiz               TEXT,
    comarca                 TEXT
)
"""

DDL_PARCELAS = """
CREATE TABLE IF NOT EXISTS parcelas (
    id              SERIAL PRIMARY KEY,
    contrato_id     INTEGER NOT NULL REFERENCES contratos(id) ON DELETE CASCADE,
    nr_parcela      INTEGER NOT NULL,
    valor_parcela   NUMERIC NOT NULL,
    data_vencimento DATE NOT NULL,
    data_pagamento  TIMESTAMP,
    pago            INTEGER DEFAULT 0,
    forma_pagamento TEXT
)
"""

DDL_PARCELAS_LIMINAR = """
CREATE TABLE IF NOT EXISTS parcelas_liminar (
    id             SERIAL PRIMARY KEY,
    contrato_id    INTEGER NOT NULL REFERENCES contratos(id) ON DELETE CASCADE,
    nr_parcela     INTEGER NOT NULL,
    valor_parcela  REAL NOT NULL,
    data_prevista  TEXT NOT NULL,
    data_pagamento TEXT,
    pago           INTEGER DEFAULT 0
)
"""

# Lista estática — nunca vem de entrada do usuário, por isso a interpolação
# direta no ALTER TABLE é segura.
COLUNAS_EXTRAS: list[tuple[str, str]] = [
    ("hon_inicial_ativo", "TEXT"),
    ("hon_inicial_valor", "REAL"),
    ("hon_inicial_parcelado", "TEXT"),
    ("hon_inicial_parcelas", "INTEGER"),
    ("hon_inicial_vlr_parcela", "REAL"),
    ("hon_liminar_fixo", "REAL"),
    ("hon_liminar_reducao_vlr", "REAL"),
    ("hon_liminar_reducao_prc", "INTEGER"),
    ("hon_exito_percentual", "REAL"),
    ("hon_exito_fixo", "REAL"),
    ("nr_processo", "TEXT"),
    ("nr_vara", "TEXT"),
    ("nome_juiz", "TEXT"),
    ("comarca", "TEXT"),
    ("tutela", "TEXT"),
    ("exito_pago", "INTEGER"),
    ("exito_data_pagamento", "TEXT"),
    ("exito_valor_recebido", "REAL"),
    # Substitui o antigo hábito de gravar "Pago" em `observacoes`, que apagava
    # as anotações do contrato quando o saldo zerava.
    ("quitado_em", "TEXT"),
]

INDICES = [
    "CREATE INDEX IF NOT EXISTS idx_parcelas_contrato ON parcelas (contrato_id)",
    "CREATE INDEX IF NOT EXISTS idx_parcelas_pendentes ON parcelas (contrato_id, nr_parcela) WHERE pago = 0",
    "CREATE INDEX IF NOT EXISTS idx_parcelas_vencimento ON parcelas (data_vencimento) WHERE pago = 0",
    "CREATE INDEX IF NOT EXISTS idx_plim_contrato ON parcelas_liminar (contrato_id)",
    "CREATE INDEX IF NOT EXISTS idx_plim_prevista ON parcelas_liminar (data_prevista) WHERE pago = 0",
    "CREATE INDEX IF NOT EXISTS idx_contratos_saldo ON contratos (saldo_devedor)",
    "CREATE INDEX IF NOT EXISTS idx_contratos_cliente ON contratos (cliente)",
]

# Aplicados separadamente: falham (e devem ser ignorados) se o banco já tiver
# parcelas duplicadas de antes.
INDICES_UNICOS = [
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_parcelas_contrato_nr ON parcelas (contrato_id, nr_parcela)",
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_plim_contrato_nr ON parcelas_liminar (contrato_id, nr_parcela)",
]


@st.cache_resource(show_spinner=False)
def inicializar_banco() -> dict[str, Any]:
    """Prepara o schema uma única vez por processo.

    Antes rodava a cada nova sessão de navegador, disparando ~20 ALTER TABLE
    por acesso.

    Importante: se a estrutura base falhar, a exceção sobe. O
    `st.cache_resource` não guarda resultado de função que levanta erro, então
    uma indisponibilidade momentânea do Supabase é reprocessada no próximo
    carregamento em vez de ficar cacheada para sempre.
    """
    with transacao() as cur:
        cur.execute(DDL_CONTRATOS)
        cur.execute(DDL_PARCELAS)
        cur.execute(DDL_PARCELAS_LIMINAR)
        for coluna, tipo in COLUNAS_EXTRAS:
            cur.execute(f"ALTER TABLE contratos ADD COLUMN IF NOT EXISTS {coluna} {tipo}")
        for indice in INDICES:
            cur.execute(indice)

    # Estes são opcionais: falham se o banco já tiver parcelas duplicadas de
    # antes. Não impedem o sistema de funcionar, então viram apenas aviso.
    avisos: list[str] = []
    for indice in INDICES_UNICOS:
        try:
            with transacao() as cur:
                cur.execute(indice)
        except Exception as erro:
            alvo = "parcelas_liminar" if "plim" in indice else "parcelas"
            avisos.append(
                f"Não foi possível criar o índice único de `{alvo}` — "
                f"provavelmente há parcelas duplicadas. Detalhe: {str(erro).strip()[:160]}"
            )

    return {"ok": not avisos, "avisos": avisos}


# =============================================================================
# 6. KEEP-ALIVE DO SUPABASE
# =============================================================================
def _ping_supabase(parametros: dict[str, Any]) -> None:
    """Abre uma conexão curta e independente do pool.

    Roda em thread de fundo, fora do contexto do Streamlit — por isso os
    parâmetros chegam prontos por argumento, sem tocar em `st.secrets`.
    """
    try:
        conexao = psycopg2.connect(**parametros)
        try:
            with conexao.cursor() as cur:
                cur.execute("SELECT 1")
        finally:
            conexao.close()
        print("[keep-alive] ping Supabase OK", flush=True)
    except Exception as erro:
        print(f"[keep-alive] erro: {erro}", flush=True)


@st.cache_resource(show_spinner=False)
def iniciar_keepalive() -> Any:
    """Um scheduler por processo.

    Antes o controle era por `st.session_state`, então cada aba aberta
    ligava um novo BackgroundScheduler e as threads iam se acumulando.
    """
    if not SCHEDULER_DISPONIVEL:
        return None
    try:
        parametros = dict(_parametros_conexao())
    except ConfiguracaoAusente:
        return None

    horas = max(1, int(_config("KEEPALIVE_HORAS", "supabase", "keepalive_horas", "6")))
    agendador = BackgroundScheduler(daemon=True)
    agendador.add_job(
        _ping_supabase,
        "interval",
        args=[parametros],
        hours=horas,
        id="keepalive_supabase",
        max_instances=1,
        coalesce=True,
        replace_existing=True,
        # Dispara já na subida: se o contêiner do Railway reiniciar, a
        # atividade volta a ser registrada na hora, e não daqui a 6 horas.
        next_run_time=agora(),
    )
    agendador.start()
    atexit.register(lambda: _silencioso(lambda: agendador.shutdown(wait=False)))
    return agendador


# =============================================================================
# 7. UTILITÁRIOS
# =============================================================================
def agora() -> datetime:
    return datetime.now(FUSO)


def hoje() -> date:
    return agora().date()


def nulo(valor: Any) -> bool:
    return not valor or str(valor).strip() in ("", "None", "nan", "NaT", "None ")


def so_digitos(valor: Any) -> str:
    return re.sub(r"\D", "", str(valor or ""))


def validar_cpf(cpf: str) -> bool:
    cpf = so_digitos(cpf)
    if len(cpf) != 11 or len(set(cpf)) == 1:
        return False
    for i in range(9, 11):
        soma = sum(int(cpf[n]) * ((i + 1) - n) for n in range(0, i))
        if (soma * 10 % 11) % 10 != int(cpf[i]):
            return False
    return True


def validar_cnpj(cnpj: str) -> bool:
    cnpj = so_digitos(cnpj)
    if len(cnpj) != 14 or len(set(cnpj)) == 1:
        return False

    def digito(n: int) -> int:
        pesos = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2] if n == 12 else [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        soma = sum(int(cnpj[i]) * pesos[i] for i in range(n)) % 11
        return 0 if soma < 2 else 11 - soma

    return digito(12) == int(cnpj[12]) and digito(13) == int(cnpj[13])


def validar_documento(documento: str) -> str | None:
    """Devolve a mensagem de erro, ou None se o documento estiver válido."""
    numero = so_digitos(documento)
    if len(numero) not in (11, 14):
        return "O documento deve ter 11 dígitos (CPF) ou 14 dígitos (CNPJ)."
    if len(numero) == 11 and not validar_cpf(numero):
        return "CPF inválido! Por favor, insira um documento real."
    if len(numero) == 14 and not validar_cnpj(numero):
        return "CNPJ inválido! Por favor, insira um documento real."
    return None


def formatar_cpf_cnpj(valor: Any) -> str:
    if nulo(valor):
        return "-"
    numero = so_digitos(valor)
    if len(numero) == 11:
        return f"{numero[:3]}.{numero[3:6]}.{numero[6:9]}-{numero[9:]}"
    if len(numero) == 14:
        return f"{numero[:2]}.{numero[2:5]}.{numero[5:8]}/{numero[8:12]}-{numero[12:]}"
    return str(valor)


def formatar_telefone(valor: Any) -> str:
    if nulo(valor):
        return "-"
    numero = so_digitos(valor)
    if len(numero) == 11:
        return f"({numero[:2]}) {numero[2:7]}-{numero[7:]}"
    if len(numero) == 10:
        return f"({numero[:2]}) {numero[2:6]}-{numero[6:]}"
    return str(valor)


def formatar_data(valor: Any) -> str:
    """Data para exibição, no padrão brasileiro (dia primeiro)."""
    if nulo(valor):
        return "-"
    texto = str(valor).strip()
    for tamanho, entrada, saida in (
        (19, "%Y-%m-%d %H:%M:%S", f"{FORMATO_DATA} %H:%M"),
        (10, "%Y-%m-%d", FORMATO_DATA),
    ):
        if len(texto) >= tamanho:
            try:
                return datetime.strptime(texto[:tamanho], entrada).strftime(saida)
            except ValueError:
                continue
    return "-"


def moeda(valor: Any) -> str:
    """Valor em real no padrão brasileiro: R$ 15.100,00.

    O format do Python só produz separador americano ("15,100.00"). O `locale`
    do sistema não é confiável em contêiner (o Railway pode não ter pt_BR
    instalado), então a troca é feita à mão: vira R$ 15.100,00 em qualquer
    servidor, sem depender de nada instalado.
    """
    try:
        bruto = f"{float(valor):,.2f}"
    except (TypeError, ValueError):
        return "R$ 0,00"
    return "R$ " + bruto.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def numero_br(valor: Any, casas: int = 2) -> str:
    """Número sem o símbolo da moeda, também no padrão brasileiro."""
    try:
        bruto = f"{float(valor):,.{casas}f}"
    except (TypeError, ValueError):
        return "0"
    return bruto.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def _cor_situacao(valor: Any) -> str:
    """CSS de fundo/texto para a célula de situação."""
    texto = str(valor)
    if "Atrasado" in texto:
        return "background-color:#FBE9E7;color:#A8201A;font-weight:600"
    if "Pendente" in texto:
        return "background-color:#FBF1D8;color:#8A5A00;font-weight:600"
    if "Pago" in texto or "Recebido" in texto:
        return "background-color:#E2F2EA;color:#10603F;font-weight:600"
    return ""


def tabela(
    df: pd.DataFrame,
    colunas_moeda: Sequence[str] = (),
    *,
    altura: int | None = None,
    ordem: Sequence[str] | None = None,
    coluna_situacao: str | None = None,
) -> None:
    """Exibe um DataFrame com moeda em pt-BR e ordenação preservada.

    O `column_config` do Streamlit só formata número com separador americano,
    e o modo "localized" depende do idioma do navegador de quem abre (num
    navegador em inglês sairia "15,100"). O Styler resolve os dois problemas:
    a formatação acontece aqui no Python, e o valor por baixo continua
    numérico — então clicar no cabeçalho ainda ordena por valor, não por texto.
    """
    if df.empty:
        st.caption("Nada a exibir.")
        return

    visao = df[list(ordem)] if ordem else df
    presentes = [c for c in colunas_moeda if c in visao.columns]
    estilo = visao.style.format({coluna: moeda for coluna in presentes})

    # Situação ganha cor de fundo: o estado é lido de relance, sem depender de
    # interpretar a bolinha do emoji. O Styler pinta a célula sem precisar de
    # HTML, que o st.dataframe não renderiza.
    if coluna_situacao and coluna_situacao in visao.columns:
        estilo = estilo.map(_cor_situacao, subset=[coluna_situacao])

    # `height` precisa ser omitido quando não definido: passar None explícito
    # levanta StreamlitInvalidHeightError.
    extras = {"height": altura} if altura else {}
    st.dataframe(estilo, use_container_width=True, hide_index=True, **extras)


def resumo_parcelamento(total: float, quantidade: int, rotulo: str = "Total") -> None:
    """Mostra a conta por extenso: total, número de parcelas e valor de cada.

    Antes aqui havia só um cartão grande com o valor de CADA parcela. O
    escritório leu esse número como se fosse o valor total e reportou que o
    sistema estava "dividindo" o valor cadastrado. O dado sempre esteve certo
    — faltava a conta estar visível por inteiro.
    """
    quantidade = max(int(quantidade), 1)
    if quantidade == 1:
        st.info(f"**{rotulo}: {moeda(total)}** — pagamento à vista, em 1 parcela.")
        return
    st.info(
        f"**{rotulo}: {moeda(total)}**  ÷  **{quantidade} parcelas**  "
        f"=  **{moeda(total / quantidade)}** por parcela"
    )


def porcentagem(fracao: float, casas: int = 1) -> str:
    """Percentual em pt-BR: 47,0% (o format do Python daria 47.0%)."""
    return numero_br(fracao * 100, casas) + "%"


def numerico(df: pd.DataFrame, *colunas: str) -> pd.DataFrame:
    """Converte colunas para número in-place (o select devolve tudo como texto)."""
    for coluna in colunas:
        if coluna in df.columns:
            df[coluna] = pd.to_numeric(df[coluna], errors="coerce").fillna(0)
    return df


def obter_status_parcela(pago: Any, vencimento: Any) -> str:
    try:
        if int(pago) == 1:
            return "🟢 Pago"
    except (TypeError, ValueError):
        pass
    try:
        data = datetime.strptime(str(vencimento)[:10], "%Y-%m-%d").date()
        if data < hoje():
            return f"🔴 Atrasado ({(hoje() - data).days} dias)"
    except (TypeError, ValueError):
        pass
    return "🟡 Pendente"


def dividir_parcelas(total: float, quantidade: int) -> list[float]:
    """Divide o valor em parcelas; a última absorve a diferença de arredondamento."""
    if quantidade <= 0:
        return []
    base = round(total / quantidade, 2)
    valores = [base] * quantidade
    valores[-1] = round(total - base * (quantidade - 1), 2)
    return valores


def gerar_vencimentos(inicio: date, quantidade: int) -> list[date]:
    base = pd.Timestamp(inicio)
    return [(base + pd.DateOffset(months=i)).date() for i in range(quantidade)]


def carimbo(data_escolhida: date) -> str:
    """Texto a gravar em data_pagamento.

    Se o recebimento é de hoje, guarda a hora real (útil no recibo). Se o
    usuário lançou uma data retroativa, guarda só a data — inventar um horário
    seria informação falsa.
    """
    if data_escolhida == hoje():
        return agora().strftime("%Y-%m-%d %H:%M:%S")
    return data_escolhida.strftime("%Y-%m-%d")


def telefone_para_banco(valor: Any) -> str | None:
    """Telefone formatado, ou None quando vazio (evita gravar '-' no banco)."""
    digitos = so_digitos(valor)
    return formatar_telefone(digitos) if digitos else None


def flash(mensagem: str, tipo: str = "success") -> None:
    """Mensagem que sobrevive ao st.rerun() — dispensa o time.sleep() de antes."""
    st.session_state["_flash"] = (tipo, mensagem)


def mostrar_flash() -> None:
    dados = st.session_state.pop("_flash", None)
    if not dados:
        return
    tipo, mensagem = dados
    getattr(st, tipo, st.info)(mensagem)


def caixa(html: str, classe: str = "caixa-cliente") -> None:
    st.markdown(f"<div class='caixa {classe}'>{html}</div>", unsafe_allow_html=True)


@st.cache_resource(show_spinner=False)
def _logo_embutido(nome: str) -> str | None:
    """Logotipo como data URI.

    Embutido em base64 porque o HTML injetado no Streamlit não enxerga
    arquivos locais, e depender do site do escritório deixaria o cabeçalho
    quebrado sempre que ele saísse do ar.
    """
    caminho = pathlib.Path(__file__).resolve().parent / "assets" / nome
    if not caminho.exists():
        return None
    return base64.b64encode(caminho.read_bytes()).decode("ascii")


def cabecalho_marca(subtitulo: str = "") -> None:
    logo = _logo_embutido("logo-color-horizontal.png")
    if logo:
        imagem = f"<img src='data:image/png;base64,{logo}' alt='{ESCRITORIO}'>"
    else:
        imagem = f"<span class='marca-titulo'>{ESCRITORIO}</span>"
    direita = (
        f"<div class='marca-titulo'>{APP_TITULO}<small>{subtitulo}</small></div>"
        if subtitulo
        else f"<div class='marca-titulo'>{APP_TITULO}</div>"
    )
    st.markdown(f"<div class='marca'>{imagem}{direita}</div>", unsafe_allow_html=True)


def linha_processo(registro: Any) -> str:
    partes = []
    for campo, rotulo in (
        ("nr_processo", "📄 Processo"),
        ("nr_vara", "🏛️ Vara"),
        ("nome_juiz", "👨‍⚖️ Juiz"),
        ("comarca", "📍 Comarca"),
    ):
        valor = registro.get(campo, "") if hasattr(registro, "get") else ""
        if not nulo(valor):
            partes.append(f"{rotulo}: {valor}")
    return " &nbsp;|&nbsp; ".join(partes)


# =============================================================================
# 8. EXPORTAÇÃO
# =============================================================================
@st.cache_data(show_spinner=False, max_entries=32)
def gerar_excel(df: pd.DataFrame, planilha: str = "Relatorio") -> bytes:
    """Excel formatado: cabeçalho destacado, filtro, colunas dimensionadas."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=planilha)
        aba = writer.sheets[planilha]

        for celula in aba[1]:
            celula.font = Font(bold=True, color="FFFFFF")
            celula.fill = PatternFill("solid", fgColor="1E4DD8")
            celula.alignment = Alignment(horizontal="center", vertical="center")

        aba.freeze_panes = "A2"
        if len(df):
            aba.auto_filter.ref = aba.dimensions

        for indice, coluna in enumerate(df.columns, start=1):
            conteudo = df[coluna].astype(str)
            largura = max([len(str(coluna))] + [len(v) for v in conteudo.head(300)] or [0])
            aba.column_dimensions[get_column_letter(indice)].width = min(max(largura + 3, 12), 45)
            if pd.api.types.is_numeric_dtype(df[coluna]):
                for celula in aba[get_column_letter(indice)][1:]:
                    celula.number_format = 'R$ #,##0.00'

    return buffer.getvalue()


def _texto_pdf(valor: Any) -> str:
    """As fontes internas do FPDF só falam latin-1 — emoji derrubava a geração."""
    return str(valor).encode("latin-1", "ignore").decode("latin-1").strip()


if PDF_DISPONIVEL:

    class _RelatorioPDF(FPDF):  # type: ignore[misc,valid-type]
        """Repete título e cabeçalho da tabela em toda página."""

        def __init__(self, titulo: str, colunas: Sequence[str], larguras: Sequence[float], **kwargs):
            super().__init__(**kwargs)
            self.titulo = titulo
            self.colunas = list(colunas)
            self.larguras = list(larguras)
            self.altura_linha = 7

        def header(self) -> None:  # noqa: D102
            self.set_font("Helvetica", "B", 14)
            self.set_text_color(20, 20, 20)
            self.cell(0, 9, _texto_pdf(self.titulo), align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_font("Helvetica", "", 8)
            self.set_text_color(110, 110, 110)
            self.cell(
                0, 5, f"Gerado em {agora().strftime(FORMATO_DATA)} {agora():%H:%M}",
                align="C", new_x="LMARGIN", new_y="NEXT",
            )
            self.ln(2)
            if self.colunas:
                self.set_fill_color(30, 77, 216)
                self.set_text_color(255, 255, 255)
                self.set_font("Helvetica", "B", 8)
                for coluna, largura in zip(self.colunas, self.larguras):
                    self.cell(largura, self.altura_linha, _texto_pdf(coluna), border=1, align="C", fill=True)
                self.ln()
            self.set_text_color(30, 30, 30)
            self.set_font("Helvetica", "", 8)

        def footer(self) -> None:  # noqa: D102
            self.set_y(-12)
            self.set_font("Helvetica", "I", 7)
            self.set_text_color(130, 130, 130)
            self.cell(0, 8, f"Página {self.page_no()} de {{nb}}", align="C")


@st.cache_data(show_spinner=False, max_entries=32)
def gerar_pdf(df: pd.DataFrame, titulo: str) -> bytes:
    """PDF paisagem com larguras de coluna proporcionais ao conteúdo."""
    if not PDF_DISPONIVEL:
        return b""

    colunas = list(df.columns)
    if df.empty or not colunas:
        pdf = _RelatorioPDF(titulo, [], [], orientation="L")
        pdf.set_auto_page_break(auto=True, margin=16)
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.set_font("Helvetica", "I", 10)
        pdf.cell(0, 10, "Nenhum registro encontrado.", align="C", new_x="LMARGIN", new_y="NEXT")
        return bytes(pdf.output())

    # Largura proporcional: nomes e observações ganham espaço, datas não.
    pesos = []
    for coluna in colunas:
        amostra = df[coluna].astype(str).head(200)
        maior = max([len(str(coluna))] + [len(v) for v in amostra] or [0])
        pesos.append(min(max(maior, 6), 40))

    pdf = _RelatorioPDF(titulo, colunas, [1] * len(colunas), orientation="L")
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.alias_nb_pages()
    largura_util = pdf.w - pdf.l_margin - pdf.r_margin
    fator = largura_util / sum(pesos)
    pdf.larguras = [peso * fator for peso in pesos]
    pdf.add_page()

    alternar = False
    for _, linha in df.iterrows():
        pdf.set_fill_color(*((240, 244, 255) if alternar else (255, 255, 255)))
        for item, largura in zip(linha, pdf.larguras):
            if nulo(item):
                texto = "-"
            elif isinstance(item, (int, float)) and not isinstance(item, bool):
                texto = moeda(item)
            else:
                texto = _texto_pdf(item)
            while pdf.get_string_width(texto) > largura - 2 and len(texto) > 3:
                texto = texto[:-4] + "..."
            pdf.cell(largura, pdf.altura_linha, texto, border=1, align="C", fill=True)
        pdf.ln()
        alternar = not alternar

    return bytes(pdf.output())


def botoes_exportacao(df: pd.DataFrame, base_nome: str, titulo: str) -> None:
    col_excel, col_pdf = st.columns(2)
    col_excel.download_button(
        "📥 Exportar para Excel",
        data=gerar_excel(df),
        file_name=f"{base_nome}_{hoje()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    if PDF_DISPONIVEL:
        col_pdf.download_button(
            "📄 Exportar para PDF",
            data=gerar_pdf(df, titulo),
            file_name=f"{base_nome}_{hoje()}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    else:
        col_pdf.caption("PDF indisponível (instale `fpdf2`).")


# =============================================================================
# 9. RECIBOS
# =============================================================================
SEPARADOR_RECIBO = "-" * 39


def montar_recibo(
    *,
    titulo: str,
    cliente: str,
    documento: Any,
    itens: Sequence[str],
    total: float,
    data: date,
    metodo: str | None = None,
    saldo_restante: float | None = None,
) -> str:
    linhas = [
        f"⚖️ *{titulo}*",
        SEPARADOR_RECIBO,
        f"*Cliente:* {cliente}",
        f"*CPF/CNPJ:* {formatar_cpf_cnpj(documento)}",
    ]
    cabecalho = f"*Data:* {data.strftime(FORMATO_DATA)}"
    if metodo:
        cabecalho += f" | *Método:* {metodo}"
    linhas += [cabecalho, SEPARADOR_RECIBO, *itens, SEPARADOR_RECIBO]
    linhas.append(f"*Total Recebido:* {moeda(total)}")
    if saldo_restante is not None:
        linhas.append(f"*Saldo Devedor Restante:* {moeda(max(saldo_restante, 0))}")
    linhas += [SEPARADOR_RECIBO, "Obrigado pela confiança!"]
    return "\n".join(linhas)


def registrar_recibo(texto: str, telefone: Any) -> None:
    st.session_state["ultimo_recibo"] = texto
    st.session_state["tel_cliente"] = str(telefone or "")


def link_whatsapp(telefone: Any, texto: str) -> str | None:
    numero = so_digitos(telefone)
    if not numero:
        return None
    # Evita virar 5555… quando o telefone já foi salvo com DDI.
    if len(numero) <= 11:
        numero = "55" + numero
    return f"https://wa.me/{numero}?text={urllib.parse.quote(texto)}"


@st.cache_data(show_spinner=False, max_entries=16)
def gerar_pdf_recibo(texto: str) -> bytes:
    if not PDF_DISPONIVEL:
        return b""
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 10, "RECIBO DE HONORARIOS", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    pdf.set_font("Helvetica", "", 11)
    for linha in texto.splitlines():
        limpa = _texto_pdf(linha.replace("*", ""))
        pdf.multi_cell(0, 7, limpa or " ", new_x="LMARGIN", new_y="NEXT")
    return bytes(pdf.output())


def painel_recibo() -> None:
    texto = st.session_state.get("ultimo_recibo")
    if not texto:
        return
    with st.container(border=True):
        st.subheader("📄 Recibo Gerado")
        st.code(texto, language="text")
        col_wpp, col_pdf, col_fechar = st.columns(3)

        url = link_whatsapp(st.session_state.get("tel_cliente"), texto)
        if url:
            col_wpp.link_button("📲 Enviar por WhatsApp", url, type="primary", use_container_width=True)
        else:
            col_wpp.caption("Cliente sem telefone cadastrado.")

        if PDF_DISPONIVEL:
            col_pdf.download_button(
                "📄 Baixar em PDF",
                data=gerar_pdf_recibo(texto),
                file_name=f"recibo_{hoje()}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

        if col_fechar.button("Limpar Tela", use_container_width=True):
            st.session_state.pop("ultimo_recibo", None)
            st.session_state.pop("tel_cliente", None)
            st.rerun()
    st.divider()


# =============================================================================
# 10. AUTENTICAÇÃO
# =============================================================================
def _bloqueio_ativo() -> int:
    ate = st.session_state.get("login_bloqueado_ate")
    if not ate:
        return 0
    restante = int((ate - agora()).total_seconds())
    return max(restante, 0)


def autenticar() -> bool:
    if st.session_state.get("autenticado"):
        return True

    _, centro, _ = st.columns([1, 2, 1])
    with centro:
        logo = _logo_embutido("logo-color-horizontal.png")
        marca = (
            f"<img src='data:image/png;base64,{logo}' alt='{ESCRITORIO}' "
            "style='max-width:290px;width:100%;margin:0 auto 6px;display:block;'>"
            if logo
            else f"<div style='font-size:1.3rem;'>{ESCRITORIO}</div>"
        )
        st.markdown(
            f"""
            <div style='text-align:center;padding:26px 0 20px;'>
                {marca}
                <div style='font-family:Fraunces,Georgia,serif;font-size:1.05rem;
                            color:#035E70;letter-spacing:.03em;'>{APP_TITULO}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        restante = _bloqueio_ativo()
        if restante:
            st.error(f"Muitas tentativas. Tente novamente em {restante}s.")
            return False

        # st.form faz o Enter enviar o login (antes só o clique funcionava).
        with st.form("form_login"):
            usuario = st.text_input("Usuário")
            senha = st.text_input("Senha", type="password")
            enviar = st.form_submit_button("Entrar", type="primary")

        if enviar:
            try:
                esperado_usuario = _config("CRED_USUARIO", "credenciais", "usuario")
                esperado_senha = _config("CRED_SENHA", "credenciais", "senha")
            except ConfiguracaoAusente as erro:
                st.error(str(erro))
                return False

            # compare_digest evita vazar o tamanho da senha por tempo de resposta.
            ok_usuario = hmac.compare_digest(usuario.strip(), esperado_usuario)
            ok_senha = hmac.compare_digest(senha, esperado_senha)
            if ok_usuario and ok_senha:
                st.session_state["autenticado"] = True
                st.session_state["usuario"] = usuario.strip()
                st.session_state.pop("tentativas_login", None)
                st.rerun()
            else:
                tentativas = st.session_state.get("tentativas_login", 0) + 1
                st.session_state["tentativas_login"] = tentativas
                if tentativas >= MAX_TENTATIVAS_LOGIN:
                    st.session_state["login_bloqueado_ate"] = agora() + timedelta(seconds=BLOQUEIO_LOGIN_SEG)
                    st.session_state["tentativas_login"] = 0
                    st.rerun()
                st.error(
                    f"Credenciais inválidas. "
                    f"({MAX_TENTATIVAS_LOGIN - tentativas} tentativa(s) restante(s))"
                )
    return False


# =============================================================================
# 11. PÁGINAS
# =============================================================================
SQL_CONTRATOS_COM_PENDENCIA = """
SELECT *
FROM contratos c
WHERE c.saldo_devedor > 0
   OR EXISTS (SELECT 1 FROM parcelas_liminar pl
               WHERE pl.contrato_id = c.id AND pl.pago = 0)
   OR (COALESCE(c.exito_pago, 0) = 0
       AND (COALESCE(c.hon_exito_percentual, 0) > 0 OR COALESCE(c.hon_exito_fixo, 0) > 0))
ORDER BY c.cliente ASC
"""

SQL_ATRASADAS = """
-- Os ::text e ::numeric não são enfeite: as colunas de data têm tipos
-- diferentes entre `parcelas` (date) e `parcelas_liminar` (text), e o UNION
-- exige que os dois lados batam. O cast normaliza sem depender do schema.
SELECT c.cliente, c.telefone, c.saldo_devedor::numeric AS saldo_devedor,
       'Honorários Iniciais' AS tipo, p.nr_parcela,
       p.valor_parcela::numeric AS valor_parcela,
       p.data_vencimento::text AS vencimento
FROM parcelas p
JOIN contratos c ON c.id = p.contrato_id
WHERE p.pago = 0 AND p.data_vencimento < %s
UNION ALL
SELECT c.cliente, c.telefone, c.saldo_devedor::numeric,
       'Liminar / Redução', pl.nr_parcela,
       pl.valor_parcela::numeric,
       pl.data_prevista::text
FROM parcelas_liminar pl
JOIN contratos c ON c.id = pl.contrato_id
WHERE pl.pago = 0 AND pl.data_prevista < %s
"""

SQL_PROXIMAS = """
SELECT c.cliente, c.telefone, 'Honorários Iniciais' AS tipo,
       p.nr_parcela, p.valor_parcela::numeric AS valor_parcela,
       p.data_vencimento::text AS vencimento
FROM parcelas p
JOIN contratos c ON c.id = p.contrato_id
WHERE p.pago = 0 AND p.data_vencimento >= %s AND p.data_vencimento <= %s
UNION ALL
SELECT c.cliente, c.telefone, 'Liminar / Redução',
       pl.nr_parcela, pl.valor_parcela::numeric,
       pl.data_prevista::text
FROM parcelas_liminar pl
JOIN contratos c ON c.id = pl.contrato_id
WHERE pl.pago = 0 AND pl.data_prevista >= %s AND pl.data_prevista <= %s
ORDER BY vencimento ASC
"""

SQL_RECEBIMENTOS_MES = """
SELECT mes, SUM(total) AS total FROM (
    SELECT LEFT(data_pagamento::text, 7) AS mes, SUM(valor_parcela::numeric) AS total
    FROM parcelas
    WHERE pago = 1 AND data_pagamento IS NOT NULL AND data_pagamento::text <> ''
    GROUP BY 1
    UNION ALL
    SELECT LEFT(data_pagamento::text, 7), SUM(valor_parcela::numeric)
    FROM parcelas_liminar
    WHERE pago = 1 AND data_pagamento IS NOT NULL AND data_pagamento::text <> ''
    GROUP BY 1
    UNION ALL
    SELECT LEFT(exito_data_pagamento::text, 7), SUM(exito_valor_recebido::numeric)
    FROM contratos
    WHERE COALESCE(exito_pago, 0) = 1
      AND exito_data_pagamento IS NOT NULL AND exito_data_pagamento::text <> ''
    GROUP BY 1
) t
WHERE mes IS NOT NULL AND mes <> ''
GROUP BY mes
ORDER BY mes
"""

SQL_IDS_ATIVOS = """
-- Contrato ativo = tem algo a receber, de qualquer natureza: saldo dos
-- honorários iniciais, parcela da redução em aberto, ou êxito acordado e
-- ainda não recebido.
SELECT c.id
FROM contratos c
WHERE COALESCE(c.saldo_devedor, 0) > 0
   OR EXISTS (SELECT 1 FROM parcelas_liminar pl
               WHERE pl.contrato_id = c.id AND pl.pago = 0)
   OR (COALESCE(c.exito_pago, 0) = 0
       AND (COALESCE(c.hon_exito_percentual, 0) > 0
            OR COALESCE(c.hon_exito_fixo, 0) > 0))
"""

SQL_ESTRUTURA = """
SELECT table_name, column_name, data_type, is_nullable, column_default
FROM information_schema.columns
WHERE table_schema = 'public'
  AND table_name IN ('contratos', 'parcelas', 'parcelas_liminar')
ORDER BY table_name, ordinal_position
"""

SQL_INSERT_CONTRATO = """
INSERT INTO contratos
    (cliente, cpf_cnpj, telefone, valor_total, saldo_devedor, data_contrato,
     observacoes, tutela,
     hon_inicial_ativo, hon_inicial_valor, hon_inicial_parcelado,
     hon_inicial_parcelas, hon_inicial_vlr_parcela,
     hon_liminar_fixo, hon_liminar_reducao_vlr, hon_liminar_reducao_prc,
     hon_exito_percentual, hon_exito_fixo,
     nr_processo, nr_vara, nome_juiz, comarca)
VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
RETURNING id
"""


def _protegido(nome: str, bloco: Callable[[], None]) -> None:
    """Roda um bloco do painel isolando a falha.

    Sem isso, uma única consulta com erro de tipo apaga a tela inteira —
    inclusive as partes que estavam funcionando.
    """
    try:
        bloco()
    except ErroBanco as erro:
        st.warning(f"Não foi possível carregar {nome}: {str(erro).strip()[:200]}", icon="⚠️")
    except Exception as erro:  # noqa: BLE001
        st.warning(f"Falha ao montar {nome}: {type(erro).__name__}: {str(erro)[:160]}", icon="⚠️")


def _mapa_contratos(df: pd.DataFrame) -> dict[str, int]:
    """Rótulos das listas de seleção: só o nome do cliente.

    O "(Contrato #12)" saiu a pedido do escritório. Ele existia por um motivo
    real, porém: o rótulo é a chave do selectbox, e dois contratos do mesmo
    cliente colidiriam — um sumiria da lista. Por isso o desempate só entra
    quando o nome de fato se repete, usando algo que o usuário reconhece
    (o processo ou a data), e o número do contrato apenas como último recurso.
    """
    nomes = df["cliente"].astype(str).str.strip()
    repetidos = set(nomes[nomes.duplicated(keep=False)])

    mapa: dict[str, int] = {}
    for _, linha in df.iterrows():
        rotulo = str(linha["cliente"]).strip()

        if rotulo in repetidos:
            processo = linha.get("nr_processo", "")
            data = linha.get("data_contrato", "")
            if not nulo(processo):
                rotulo = f"{rotulo} — {processo}"
            elif not nulo(data):
                rotulo = f"{rotulo} — {formatar_data(data)}"
            else:
                rotulo = f"{rotulo} — #{linha['id']}"

        # Rede de segurança: mesmo nome, mesmo processo, mesma data.
        base, tentativa = rotulo, 2
        while rotulo in mapa:
            rotulo = f"{base} ({tentativa})"
            tentativa += 1

        mapa[rotulo] = int(linha["id"])
    return mapa


# ---------------------------------------------------------------- DASHBOARD --
def pagina_dashboard() -> None:
    st.header("Resumo Financeiro")

    df_contratos = select_db("SELECT * FROM contratos ORDER BY cliente ASC")
    if df_contratos.empty:
        st.info("Nenhum contrato registrado.")
        return

    numerico(df_contratos, "valor_total", "saldo_devedor")
    # "Ativo" era só quem devia honorários iniciais. Para este escritório isso
    # dava quase sempre zero: boa parte dos contratos não tem cobrança inicial
    # e vive da redução da liminar e do êxito. Um contrato com parcela de
    # redução a vencer está ativo, mesmo com saldo inicial zerado.
    ids_ativos = select_db(SQL_IDS_ATIVOS)
    if ids_ativos.empty:
        df_ativos = df_contratos.iloc[0:0].copy()
    else:
        df_ativos = df_contratos[
            df_contratos["id"].isin(set(ids_ativos["id"].astype(int)))
        ].copy()

    df_liminar = select_db(
        """
        SELECT COALESCE(SUM(CASE WHEN pago = 1 THEN valor_parcela END), 0) AS recebido,
               COALESCE(SUM(CASE WHEN pago = 0 THEN valor_parcela END), 0) AS pendente
        FROM parcelas_liminar
        """
    )
    liminar_pendente = float(df_liminar.iloc[0]["pendente"]) if not df_liminar.empty else 0.0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Honorários Contratados", moeda(df_contratos["valor_total"].sum()))
    m2.metric("Saldo Devedor Total", moeda(df_contratos["saldo_devedor"].sum()))
    m3.metric("Contratos Ativos", len(df_ativos))
    m4.metric("Redução a Receber", moeda(liminar_pendente))
    st.divider()

    # Cada bloco é isolado: uma consulta com problema mostra um aviso no lugar
    # dela em vez de derrubar o painel inteiro, como acontecia antes.
    _protegido("alerta de inadimplência", _bloco_inadimplencia)
    _protegido("próximos vencimentos", _bloco_proximos_vencimentos)

    if df_ativos.empty:
        st.success("Nenhum contrato com saldo devedor em aberto.")
        _protegido("gráfico de recebimentos", _grafico_recebimentos)
        return

    col_titulo, col_atalho = st.columns([2, 1])
    col_titulo.subheader("Contratos Ativos (Em Aberto)")
    with col_atalho:
        st.markdown("⚡ **Atalho Rápido:**")
        mapa = _mapa_contratos(df_ativos)
        escolhido = st.selectbox(
            "Selecione o cliente:", options=list(mapa.keys()), label_visibility="collapsed"
        )

        def ir_para_pagamentos() -> None:
            st.session_state["cliente_foco"] = mapa[escolhido]
            st.session_state["rad_nav"] = "💰 Pagamentos"

        st.button("Ir para Pagamento ➡", type="primary", on_click=ir_para_pagamentos)

    df_visao = pd.DataFrame(
        {
            "Cliente": df_ativos["cliente"],
            "CPF/CNPJ": df_ativos["cpf_cnpj"].apply(formatar_cpf_cnpj),
            "Telefone": df_ativos["telefone"].apply(formatar_telefone),
            "Data Contrato": df_ativos["data_contrato"].apply(formatar_data),
            "Valor Total": df_ativos["valor_total"],
            "Saldo Pendente": df_ativos["saldo_devedor"],
            "Observações": df_ativos["observacoes"].apply(lambda v: "-" if nulo(v) else str(v)),
        }
    )
    tabela(df_visao, ["Valor Total", "Saldo Pendente"])
    st.divider()
    botoes_exportacao(df_visao, "contratos_ativos", "Relatório de Contratos Ativos")
    _protegido("gráfico de recebimentos", _grafico_recebimentos)


def _bloco_inadimplencia() -> None:
    referencia = hoje().isoformat()
    # O filtro de vencidos passou para o SQL: antes o sistema trazia TODAS as
    # parcelas em aberto do banco e descartava a maioria no pandas.
    df = select_db(SQL_ATRASADAS, (referencia, referencia))
    if df.empty:
        return

    numerico(df, "valor_parcela", "saldo_devedor")
    df["vencimento_dt"] = pd.to_datetime(df["vencimento"], errors="coerce")
    df = df.dropna(subset=["vencimento_dt"])
    if df.empty:
        return

    df["dias"] = (pd.Timestamp(hoje()) - df["vencimento_dt"]).dt.days
    resumo = (
        df.groupby(["cliente", "telefone"])
        .agg(
            qtd=("nr_parcela", "count"),
            valor=("valor_parcela", "sum"),
            pior=("dias", "max"),
            tipos=("tipo", lambda coluna: " + ".join(sorted(set(coluna)))),
        )
        .reset_index()
        .sort_values("pior", ascending=False)
    )

    st.error(
        f"🚨 Atenção: {len(resumo)} cliente(s) em inadimplência! "
        f"Total atrasado: **{moeda(resumo['valor'].sum())}**"
    )
    resumo["telefone"] = resumo["telefone"].apply(formatar_telefone)
    resumo.columns = [
        "Cliente", "Telefone", "Parcelas Atrasadas",
        "Valor Atrasado", "Dias do Pior Atraso", "Origem",
    ]
    tabela(resumo, ["Valor Atrasado"])
    st.divider()


def _bloco_proximos_vencimentos(dias: int = 15) -> None:
    inicio = hoje().isoformat()
    fim = (hoje() + timedelta(days=dias)).isoformat()
    df = select_db(SQL_PROXIMAS, (inicio, fim, inicio, fim))
    if df.empty:
        return

    numerico(df, "valor_parcela")
    with st.expander(f"📆 Vencem nos próximos {dias} dias ({len(df)})", expanded=False):
        visao = pd.DataFrame(
            {
                "Cliente": df["cliente"],
                "Telefone": df["telefone"].apply(formatar_telefone),
                "Origem": df["tipo"],
                "Parcela": df["nr_parcela"],
                "Valor": df["valor_parcela"],
                "Vencimento": df["vencimento"].apply(formatar_data),
            }
        )
        tabela(visao, ["Valor"])
        st.caption(f"Total previsto: **{moeda(df['valor_parcela'].sum())}**")


def _grafico_recebimentos() -> None:
    try:
        df = select_db(SQL_RECEBIMENTOS_MES)
    except ErroBanco as erro:
        # Antes esta falha era engolida em silêncio: o gráfico simplesmente
        # não aparecia e não havia como saber por quê.
        st.divider()
        st.caption(f"📈 Gráfico de recebimentos indisponível: {str(erro).strip()[:180]}")
        return
    if df.empty:
        return

    numerico(df, "total")
    df = df.tail(12).copy()
    df["Mês"] = df["mes"].apply(
        lambda valor: f"{str(valor)[5:7]}/{str(valor)[:4]}" if len(str(valor)) >= 7 else str(valor)
    )
    st.divider()
    st.subheader("📈 Recebimentos por Mês")

    # st.bar_chart rotula o eixo em formato americano ("15,000"). O Altair
    # aceita o locale do Vega, então o eixo e a dica saem em pt-BR.
    grafico = (
        alt.Chart(df)
        .mark_bar(color="#035E70", cornerRadiusTopLeft=3, cornerRadiusTopRight=3)
        .encode(
            x=alt.X("Mês:N", title=None, sort=None, axis=alt.Axis(labelAngle=0)),
            y=alt.Y("total:Q", title=None, axis=alt.Axis(format="$,.0f")),
            tooltip=[
                alt.Tooltip("Mês:N", title="Mês"),
                alt.Tooltip("total:Q", title="Recebido", format="$,.2f"),
            ],
        )
        .properties(height=260)
        .configure(locale=LOCALE_VEGA)
        .configure_axis(grid=True, gridColor="#E7F0F1", domainColor="#C9D6D8", labelColor="#43545A")
        .configure_view(strokeWidth=0)
    )
    st.altair_chart(grafico, use_container_width=True)
    st.caption(f"Total recebido no período exibido: **{moeda(df['total'].sum())}**")


# ----------------------------------------------------------- NOVO CONTRATO --
def pagina_novo_contrato() -> None:
    st.header("Cadastrar Novo Contrato")

    st.subheader("📋 Informações Básicas")
    col1, col2 = st.columns(2)
    nome = col1.text_input("Nome do Cliente")
    documento = col2.text_input("CPF ou CNPJ (somente números)", placeholder="Ex: 00000000000")
    telefone = col1.text_input("Telefone (somente números)", placeholder="Ex: 11999998888")
    data_contrato = col2.date_input("Data do Contrato", value=hoje(), format=FORMATO_DATA_WIDGET)

    st.divider()

    st.subheader("💰 Honorários Iniciais")
    col1, col2 = st.columns(2)
    ini_ativo = col1.selectbox("Há cobrança inicial?", ["Não", "Sim"], key="novo_ini_ativo")
    cobra_inicial = ini_ativo == "Sim"
    ini_valor = col2.number_input(
        "Valor Total (R$)",
        min_value=0.0, step=100.0, format="%.2f",
        key="novo_ini_valor", disabled=not cobra_inicial,
    )
    if not cobra_inicial:
        ini_valor = 0.0

    ini_parcelado = col1.selectbox(
        "Pagamento parcelado?", ["Não", "Sim"],
        key="novo_ini_parcelado", disabled=not cobra_inicial,
    )

    # Fonte única de verdade: antes existiam DOIS controles de parcelamento
    # (o "Nº de Parcelas" e o selectbox de "controle"), que podiam divergir e
    # gravavam quantidades diferentes no contrato e nas parcelas.
    quantidade = 1
    primeiro_vencimento = data_contrato
    if cobra_inicial and ini_valor > 0:
        col3, col4, col5 = st.columns(3)
        if ini_parcelado == "Sim":
            quantidade = int(
                col3.number_input(
                    "Quantidade de Parcelas", min_value=1, max_value=60, value=1, step=1, key="novo_ini_qtd"
                )
            )
        else:
            col3.metric("Forma", "À vista")
        primeiro_vencimento = col4.date_input(
            "Vencimento da 1ª Parcela", value=data_contrato, key="novo_ini_venc", format=FORMATO_DATA_WIDGET)
        col5.write("")
        resumo_parcelamento(ini_valor, quantidade, "Honorários iniciais")

    valor_parcela = round(ini_valor / quantidade, 2) if quantidade > 0 and ini_valor > 0 else 0.0

    st.divider()

    st.subheader("⚖️ Honorários da Liminar")
    col1, col2 = st.columns(2)
    tutela = col1.selectbox("Status da Tutela", STATUS_TUTELA)
    liminar_fixo = col2.number_input(
        "Honorários Fixos da Liminar (R$)", min_value=0.0, step=100.0, format="%.2f", key="novo_lim_fixo"
    )
    col3, col4 = st.columns(2)
    reducao_valor = col3.number_input(
        "Valor Efetivo da Redução Obtida (R$)", min_value=0.0, step=100.0, format="%.2f", key="novo_red_vlr"
    )
    reducao_parcelas = int(
        col4.number_input(
            "Nº de Parcelas da Redução", min_value=0, max_value=360, value=0, step=1, key="novo_red_prc"
        )
    )

    st.divider()

    st.subheader("🏆 Honorários de Êxito")
    col1, col2 = st.columns(2)
    exito_pct = col1.number_input(
        "Percentual de Êxito (%)", min_value=0.0, max_value=100.0, step=0.5, format="%.2f", key="novo_ex_pct"
    )
    exito_fixo = col2.number_input(
        "Valor Fixo de Êxito (R$)", min_value=0.0, step=100.0, format="%.2f", key="novo_ex_fixo"
    )

    st.divider()

    st.subheader("📁 Dados do Processo")
    col1, col2 = st.columns(2)
    nr_processo = col1.text_input("Número do Processo", placeholder="Ex: 0000000-00.0000.0.00.0000")
    nr_vara = col2.text_input("Número da Vara", placeholder="Ex: 3ª Vara Cível")
    col3, col4 = st.columns(2)
    nome_juiz = col3.text_input("Nome do Juiz")
    comarca = col4.text_input("Comarca")
    observacoes = st.text_area("Observações (anotações extras)")

    st.divider()

    if not st.button("Salvar Contrato", type="primary"):
        return

    if not nome.strip():
        st.error("Nome obrigatório.")
        return
    erro_documento = validar_documento(documento)
    if erro_documento:
        st.error(erro_documento)
        return

    documento_limpo = so_digitos(documento)
    duplicado = select_db(
        "SELECT cliente FROM contratos WHERE regexp_replace(COALESCE(cpf_cnpj,''), '\\D', '', 'g') = %s LIMIT 1",
        (documento_limpo,),
        cache=False,
    )
    if not duplicado.empty:
        st.warning(
            f"⚠️ Já existe contrato para este documento (**{duplicado.iloc[0]['cliente']}**). "
            "O novo contrato será criado assim mesmo."
        )

    valores = dividir_parcelas(ini_valor, quantidade) if ini_valor > 0 else []
    vencimentos = gerar_vencimentos(primeiro_vencimento, quantidade) if valores else []

    with transacao() as cur:
        cur.execute(
            SQL_INSERT_CONTRATO,
            (
                nome.strip(),
                formatar_cpf_cnpj(documento_limpo),
                telefone_para_banco(telefone),
                ini_valor,
                ini_valor,
                data_contrato.strftime("%Y-%m-%d"),
                observacoes.strip() or None,
                tutela,
                ini_ativo,
                ini_valor if cobra_inicial else None,
                ini_parcelado,
                quantidade if ini_parcelado == "Sim" else None,
                valor_parcela if ini_parcelado == "Sim" else None,
                liminar_fixo or None,
                reducao_valor or None,
                reducao_parcelas or None,
                exito_pct or None,
                exito_fixo or None,
                nr_processo.strip() or None,
                nr_vara.strip() or None,
                nome_juiz.strip() or None,
                comarca.strip() or None,
            ),
        )
        contrato_id = cur.fetchone()[0]

        if valores:
            # execute_values manda tudo numa ida só ao banco (antes era um
            # INSERT + commit por parcela).
            execute_values(
                cur,
                "INSERT INTO parcelas (contrato_id, nr_parcela, valor_parcela, data_vencimento) VALUES %s",
                [
                    (contrato_id, numero, valor, vencimento.strftime("%Y-%m-%d"))
                    for numero, (valor, vencimento) in enumerate(zip(valores, vencimentos), start=1)
                ],
            )

    flash(f"Contrato #{contrato_id} cadastrado com sucesso!")
    st.rerun()


# --------------------------------------------------------------- PAGAMENTOS --
def pagina_pagamentos() -> None:
    st.header("Registrar Recebimento")
    painel_recibo()

    col_sel, col_todos = st.columns([3, 1])
    with col_todos:
        st.write("")
        mostrar_todos = st.checkbox("Mostrar todos", value=False, help="Inclui contratos já quitados")
    consulta = (
        "SELECT * FROM contratos ORDER BY cliente ASC" if mostrar_todos else SQL_CONTRATOS_COM_PENDENCIA
    )
    df_contratos = select_db(consulta)

    if df_contratos.empty:
        st.info("Não há contratos pendentes.")
        return

    numerico(df_contratos, "valor_total", "saldo_devedor")
    mapa = _mapa_contratos(df_contratos)
    rotulos = list(mapa.keys())

    if "cliente_foco" in st.session_state:
        foco = st.session_state.pop("cliente_foco")
        for rotulo in rotulos:
            if mapa[rotulo] == foco:
                st.session_state["select_cliente"] = rotulo
                break

    if st.session_state.get("select_cliente") not in rotulos:
        st.session_state.pop("select_cliente", None)

    with col_sel:
        rotulo_sel = st.selectbox("Selecione o Cliente", options=rotulos, key="select_cliente")
    contrato_id = mapa[rotulo_sel]
    contrato = df_contratos[df_contratos["id"] == contrato_id].iloc[0]

    valor_total = float(contrato["valor_total"])
    saldo = float(contrato["saldo_devedor"])
    progresso = max(0.0, min(1.0, (valor_total - saldo) / valor_total)) if valor_total > 0 else 0.0

    # Dados do cliente e situação lado a lado: antes eram quatro blocos
    # empilhados que empurravam a ação para fora da tela.
    col_dados, col_valores = st.columns([3, 2])
    with col_dados:
        processo = linha_processo(contrato)
        caixa(
            f"<b>{contrato['cliente']}</b><br>"
            f"💳 {formatar_cpf_cnpj(contrato['cpf_cnpj'])} &nbsp;|&nbsp; "
            f"📞 {formatar_telefone(contrato['telefone'])}"
            + (f"<br><span style='font-size:.88rem;color:#43545A;'>{processo}</span>" if processo else "")
        )
    with col_valores:
        m1, m2 = st.columns(2)
        m1.metric("Valor Total", moeda(valor_total))
        m2.metric("Saldo Restante", moeda(saldo))
        st.progress(progresso, text=f"Pago: {porcentagem(progresso)}")

    if not nulo(contrato["observacoes"]):
        caixa(f"<b>Notas:</b> {contrato['observacoes']}", "caixa-nota")

    aba_inicial, aba_liminar, aba_exito, aba_combinado = st.tabs(
        ["💰 Honorários Iniciais", "⚖️ Liminar / Redução", "🏆 Êxito", "📋 Combinado"]
    )
    with aba_inicial:
        _tab_honorarios_iniciais(contrato_id, contrato)
    with aba_liminar:
        _tab_liminar(contrato_id, contrato)
    with aba_exito:
        _tab_exito(contrato_id, contrato)
    with aba_combinado:
        _tab_combinado(contrato_id, contrato, saldo)


def _tabela_parcelas(df: pd.DataFrame, coluna_data: str, rotulo_data: str) -> None:
    visao = pd.DataFrame(
        {
            "Parcela": df["nr_parcela"],
            "Valor": df["valor_parcela"],
            rotulo_data: df[coluna_data].apply(formatar_data),
            "Pago em": df["data_pagamento"].apply(formatar_data),
            "Status": df.apply(lambda linha: obter_status_parcela(linha["pago"], linha[coluna_data]), axis=1),
        }
    )
    if "forma_pagamento" in df.columns:
        visao["Método"] = df["forma_pagamento"].apply(lambda v: "-" if nulo(v) else str(v))
    tabela(visao, ["Valor"], coluna_situacao="Status")


def _tab_honorarios_iniciais(contrato_id: int, contrato: Any) -> None:
    df = select_db(
        "SELECT * FROM parcelas WHERE contrato_id = %s ORDER BY nr_parcela", (contrato_id,)
    )
    if df.empty:
        st.info("Nenhuma parcela de honorários iniciais cadastrada.")
        return

    numerico(df, "pago", "valor_parcela")
    df["pago"] = df["pago"].astype(int)

    pendentes = df[df["pago"] == 0]
    if pendentes.empty:
        st.success("🎉 Todas as parcelas iniciais já foram pagas!")
        _tabela_parcelas(df, "data_vencimento", "Vencimento")
        return

    # Formulário antes da tabela: a ação principal passa a caber na tela sem
    # rolar. A lista completa fica logo abaixo, como consulta.
    with st.container(border=True):
        col1, col2 = st.columns(2)
        opcoes = {
            f"Parcela {linha.nr_parcela} — vence {formatar_data(linha.data_vencimento)}": int(linha.nr_parcela)
            for linha in pendentes.itertuples()
        }
        rotulo = col1.selectbox("Qual parcela pagar?", list(opcoes), key=f"ini_parc_{contrato_id}")
        numero = opcoes[rotulo]
        sugerido = float(pendentes[pendentes["nr_parcela"] == numero]["valor_parcela"].iloc[0])
        valor_pago = col2.number_input(
            "Valor Recebido (R$)", value=sugerido, min_value=0.0, format="%.2f",
            key=f"ini_vlr_{contrato_id}",
        )
        col3, col4 = st.columns(2)
        forma = col3.selectbox("Método de Recebimento", FORMAS_PAGAMENTO, key=f"ini_forma_{contrato_id}")
        data_pagamento = col4.date_input(
            "Data do Recebimento", value=hoje(), key=f"ini_data_{contrato_id}",
            format=FORMATO_DATA_WIDGET,
        )
        confirmar = st.button("Confirmar Pagamento", type="primary", key=f"ini_btn_{contrato_id}")

    st.caption("Todas as parcelas deste contrato:")
    _tabela_parcelas(df, "data_vencimento", "Vencimento")

    if not confirmar:
        return

    novo_saldo, quitou = _baixar_parcela_inicial(contrato_id, numero, valor_pago, forma, data_pagamento)
    registrar_recibo(
        montar_recibo(
            titulo="RECIBO DE HONORÁRIOS",
            cliente=contrato["cliente"],
            documento=contrato["cpf_cnpj"],
            itens=[f"💰 Honorários Iniciais — Parcela {numero}: {moeda(valor_pago)}"],
            total=valor_pago,
            data=data_pagamento,
            metodo=forma,
            saldo_restante=novo_saldo,
        ),
        contrato["telefone"],
    )
    if quitou:
        st.balloons()
    flash(f"Parcela {numero} baixada com sucesso!")
    st.rerun()


def _baixar_parcela_inicial(
    contrato_id: int, numero: int, valor: float, forma: str, data_pagamento: date
) -> tuple[float, bool]:
    """Baixa a parcela e atualiza o saldo numa única transação.

    O saldo é decrementado pelo próprio banco (`saldo_devedor - valor`) em vez
    de ser recalculado a partir do valor lido na tela, que podia estar
    desatualizado se outra aba tivesse registrado um pagamento no meio.
    """
    with transacao() as cur:
        cur.execute(
            """UPDATE parcelas
                  SET pago = 1, data_pagamento = %s, forma_pagamento = %s
                WHERE contrato_id = %s AND nr_parcela = %s AND pago = 0""",
            (carimbo(data_pagamento), forma, contrato_id, numero),
        )
        if cur.rowcount == 0:
            raise ErroBanco("Esta parcela já havia sido baixada. A tela foi atualizada.")

        cur.execute(
            """UPDATE contratos
                  SET saldo_devedor = GREATEST(COALESCE(saldo_devedor, 0) - %s, 0)
                WHERE id = %s
            RETURNING saldo_devedor""",
            (valor, contrato_id),
        )
        novo_saldo = float(cur.fetchone()[0] or 0)

        # Marca a quitação sem sobrescrever as observações do contrato.
        cur.execute(
            """UPDATE contratos
                  SET quitado_em = CASE WHEN saldo_devedor <= 0
                                        THEN COALESCE(quitado_em, %s) END
                WHERE id = %s""",
            (data_pagamento.strftime("%Y-%m-%d"), contrato_id),
        )
    return novo_saldo, novo_saldo <= 0


def _tab_liminar(contrato_id: int, contrato: Any) -> None:
    df = select_db(
        "SELECT * FROM parcelas_liminar WHERE contrato_id = %s ORDER BY nr_parcela", (contrato_id,)
    )
    tutela = str(contrato.get("tutela", "") or "Pendente")

    if df.empty:
        if tutela in TUTELA_COM_REDUCAO:
            st.info(
                "Tutela deferida, mas as parcelas da redução ainda não foram cadastradas. "
                "Acesse **📂 Meus Contratos** para criar as parcelas."
            )
        else:
            st.info(
                f"Status da tutela: **{tutela}**. "
                "Quando a tutela for deferida, cadastre as parcelas em **📂 Meus Contratos**."
            )
        return

    numerico(df, "pago", "valor_parcela")
    df["pago"] = df["pago"].astype(int)
    _tabela_parcelas(df, "data_prevista", "Previsão")

    pendentes = df[df["pago"] == 0]
    if pendentes.empty:
        st.success("🎉 Todas as parcelas da redução já foram recebidas!")
        return

    col1, col2 = st.columns(2)
    opcoes = {
        f"Parcela {linha.nr_parcela} — Prev. {formatar_data(linha.data_prevista)} — {moeda(linha.valor_parcela)}": int(
            linha.nr_parcela
        )
        for linha in pendentes.itertuples()
    }
    rotulo = col1.selectbox("Qual parcela recebeu?", list(opcoes), key=f"lim_sel_{contrato_id}")
    numero = opcoes[rotulo]
    sugerido = float(pendentes[pendentes["nr_parcela"] == numero]["valor_parcela"].iloc[0])
    valor_recebido = col2.number_input(
        "Valor Recebido (R$)", value=sugerido, min_value=0.0, format="%.2f", key=f"lim_vlr_{contrato_id}"
    )
    data_recebimento = st.date_input("Data do Recebimento", value=hoje(), key=f"lim_data_{contrato_id}", format=FORMATO_DATA_WIDGET)

    if not st.button("✅ Confirmar Recebimento", type="primary", key=f"lim_btn_{contrato_id}"):
        return

    _baixar_parcela_liminar(contrato_id, numero, data_recebimento)
    registrar_recibo(
        montar_recibo(
            titulo="RECIBO — LIMINAR / REDUÇÃO",
            cliente=contrato["cliente"],
            documento=contrato["cpf_cnpj"],
            itens=[f"⚖️ Redução da Liminar — Parcela {numero}: {moeda(valor_recebido)}"],
            total=valor_recebido,
            data=data_recebimento,
        ),
        contrato["telefone"],
    )
    flash(f"Parcela {numero} da redução registrada!")
    st.rerun()


def _baixar_parcela_liminar(contrato_id: int, numero: int, data_recebimento: date) -> None:
    with transacao() as cur:
        cur.execute(
            """UPDATE parcelas_liminar
                  SET pago = 1, data_pagamento = %s
                WHERE contrato_id = %s AND nr_parcela = %s AND pago = 0""",
            (data_recebimento.strftime("%Y-%m-%d"), contrato_id, numero),
        )
        if cur.rowcount == 0:
            raise ErroBanco("Esta parcela já havia sido registrada. A tela foi atualizada.")


def _tab_exito(contrato_id: int, contrato: Any) -> None:
    ja_pago = int(contrato.get("exito_pago") or 0)
    percentual = float(contrato.get("hon_exito_percentual") or 0)
    fixo = float(contrato.get("hon_exito_fixo") or 0)

    if percentual == 0 and fixo == 0:
        st.info(
            "Nenhum honorário de êxito configurado para este contrato. "
            "Configure em **📂 Meus Contratos → Editar Contrato**."
        )
        return

    if ja_pago == 1:
        recebido = float(contrato.get("exito_valor_recebido") or 0)
        data_recebido = formatar_data(contrato.get("exito_data_pagamento"))
        st.success(f"🏆 Honorários de êxito já recebidos em **{data_recebido}**: **{moeda(recebido)}**")
        if st.button("↩️ Estornar recebimento de êxito", key=f"exit_estorno_{contrato_id}"):
            exec_db(
                """UPDATE contratos
                      SET exito_pago = 0, exito_data_pagamento = NULL, exito_valor_recebido = NULL
                    WHERE id = %s""",
                (contrato_id,),
            )
            flash("Recebimento de êxito estornado.", "warning")
            st.rerun()
        return

    if percentual > 0:
        st.info(f"Percentual de êxito acordado: **{numero_br(percentual)}%** sobre o valor da causa.")
    if fixo > 0:
        st.info(f"Valor fixo de êxito acordado: **{moeda(fixo)}**")

    col1, col2 = st.columns(2)
    valor_recebido = col1.number_input(
        "Valor Recebido (R$)",
        min_value=0.01, step=100.0, format="%.2f",
        value=fixo if fixo > 0 else 100.0,
        key=f"exit_vlr_{contrato_id}",
    )
    data_recebimento = col2.date_input("Data do Recebimento", value=hoje(), key=f"exit_data_{contrato_id}", format=FORMATO_DATA_WIDGET)

    if not st.button("🏆 Confirmar Recebimento de Êxito", type="primary", key=f"exit_btn_{contrato_id}"):
        return

    exec_db(
        """UPDATE contratos
              SET exito_pago = 1, exito_data_pagamento = %s, exito_valor_recebido = %s
            WHERE id = %s""",
        (data_recebimento.strftime("%Y-%m-%d"), valor_recebido, contrato_id),
    )
    registrar_recibo(
        montar_recibo(
            titulo="RECIBO — HONORÁRIOS DE ÊXITO",
            cliente=contrato["cliente"],
            documento=contrato["cpf_cnpj"],
            itens=[f"🏆 Honorários de Êxito: {moeda(valor_recebido)}"],
            total=valor_recebido,
            data=data_recebimento,
        ),
        contrato["telefone"],
    )
    flash("Honorários de êxito registrados!")
    st.rerun()


def _tab_combinado(contrato_id: int, contrato: Any, saldo: float) -> None:
    st.markdown("Registre múltiplos recebimentos de uma vez (ex: inicial + liminar).")

    df_iniciais = select_db(
        "SELECT * FROM parcelas WHERE contrato_id = %s AND pago = 0 ORDER BY nr_parcela", (contrato_id,)
    )
    df_liminar = select_db(
        "SELECT * FROM parcelas_liminar WHERE contrato_id = %s AND pago = 0 ORDER BY nr_parcela",
        (contrato_id,),
    )
    exito_pendente = int(contrato.get("exito_pago") or 0) == 0 and (
        float(contrato.get("hon_exito_percentual") or 0) > 0
        or float(contrato.get("hon_exito_fixo") or 0) > 0
    )
    exito_fixo = float(contrato.get("hon_exito_fixo") or 0)

    if df_iniciais.empty and df_liminar.empty and not exito_pendente:
        st.success("Não há pendências a registrar para este contrato.")
        return

    marcar_inicial = st.checkbox(
        "💰 Honorários Iniciais", key=f"comb_ini_{contrato_id}", disabled=df_iniciais.empty
    )
    marcar_liminar = st.checkbox(
        "⚖️ Liminar / Redução", key=f"comb_lim_{contrato_id}", disabled=df_liminar.empty
    )
    marcar_exito = st.checkbox("🏆 Êxito", key=f"comb_exit_{contrato_id}", disabled=not exito_pendente)

    parcela_inicial: int | None = None
    parcela_liminar: int | None = None
    valor_inicial = valor_liminar = valor_exito = 0.0

    if marcar_inicial and not df_iniciais.empty:
        numerico(df_iniciais, "valor_parcela")
        opcoes = {
            f"Parc {linha.nr_parcela} (Venc: {formatar_data(linha.data_vencimento)})": int(linha.nr_parcela)
            for linha in df_iniciais.itertuples()
        }
        col1, col2 = st.columns(2)
        parcela_inicial = opcoes[col1.selectbox("Parcela inicial:", list(opcoes), key=f"comb_ini_sel_{contrato_id}")]
        valor_inicial = col2.number_input(
            "Valor inicial (R$)",
            value=float(df_iniciais[df_iniciais["nr_parcela"] == parcela_inicial]["valor_parcela"].iloc[0]),
            min_value=0.0, format="%.2f", key=f"comb_ini_vlr_{contrato_id}",
        )

    if marcar_liminar and not df_liminar.empty:
        numerico(df_liminar, "valor_parcela")
        opcoes = {
            f"Parcela {linha.nr_parcela} — {moeda(linha.valor_parcela)}": int(linha.nr_parcela)
            for linha in df_liminar.itertuples()
        }
        col3, col4 = st.columns(2)
        parcela_liminar = opcoes[col3.selectbox("Parcela liminar:", list(opcoes), key=f"comb_lim_sel_{contrato_id}")]
        valor_liminar = col4.number_input(
            "Valor liminar (R$)",
            value=float(df_liminar[df_liminar["nr_parcela"] == parcela_liminar]["valor_parcela"].iloc[0]),
            min_value=0.0, format="%.2f", key=f"comb_lim_vlr_{contrato_id}",
        )

    if marcar_exito and exito_pendente:
        valor_exito = st.number_input(
            "Valor Êxito (R$)",
            value=exito_fixo if exito_fixo > 0 else 100.0,
            min_value=0.0, format="%.2f", key=f"comb_exit_vlr_{contrato_id}",
        )

    if not (marcar_inicial or marcar_liminar or marcar_exito):
        return

    forma = st.selectbox("Método de Recebimento", FORMAS_PAGAMENTO, key=f"comb_forma_{contrato_id}")
    data_recebimento = st.date_input("Data do Recebimento", value=hoje(), key=f"comb_data_{contrato_id}", format=FORMATO_DATA_WIDGET)
    total = valor_inicial + valor_liminar + valor_exito
    st.metric("Total a Registrar", moeda(total))

    if not st.button("✅ Confirmar Pagamento Combinado", type="primary", key=f"comb_btn_{contrato_id}"):
        return

    itens: list[str] = []
    novo_saldo: float | None = None

    # Tudo dentro de UMA transação: ou os três recebimentos entram, ou nenhum.
    with transacao() as cur:
        if marcar_inicial and parcela_inicial is not None:
            cur.execute(
                """UPDATE parcelas
                      SET pago = 1, data_pagamento = %s, forma_pagamento = %s
                    WHERE contrato_id = %s AND nr_parcela = %s AND pago = 0""",
                (data_recebimento.strftime("%Y-%m-%d"), forma, contrato_id, parcela_inicial),
            )
            if cur.rowcount == 0:
                raise ErroBanco("A parcela inicial selecionada já havia sido baixada.")
            cur.execute(
                """UPDATE contratos
                      SET saldo_devedor = GREATEST(COALESCE(saldo_devedor, 0) - %s, 0)
                    WHERE id = %s
                RETURNING saldo_devedor""",
                (valor_inicial, contrato_id),
            )
            novo_saldo = float(cur.fetchone()[0] or 0)
            cur.execute(
                """UPDATE contratos
                      SET quitado_em = CASE WHEN saldo_devedor <= 0
                                            THEN COALESCE(quitado_em, %s) END
                    WHERE id = %s""",
                (data_recebimento.strftime("%Y-%m-%d"), contrato_id),
            )
            itens.append(f"💰 Honorários Iniciais (Parc. {parcela_inicial}): {moeda(valor_inicial)}")

        if marcar_liminar and parcela_liminar is not None:
            cur.execute(
                """UPDATE parcelas_liminar
                      SET pago = 1, data_pagamento = %s
                    WHERE contrato_id = %s AND nr_parcela = %s AND pago = 0""",
                (data_recebimento.strftime("%Y-%m-%d"), contrato_id, parcela_liminar),
            )
            if cur.rowcount == 0:
                raise ErroBanco("A parcela da liminar selecionada já havia sido registrada.")
            itens.append(f"⚖️ Liminar / Redução (Parc. {parcela_liminar}): {moeda(valor_liminar)}")

        if marcar_exito and exito_pendente:
            cur.execute(
                """UPDATE contratos
                      SET exito_pago = 1, exito_data_pagamento = %s, exito_valor_recebido = %s
                    WHERE id = %s AND COALESCE(exito_pago, 0) = 0""",
                (data_recebimento.strftime("%Y-%m-%d"), valor_exito, contrato_id),
            )
            itens.append(f"🏆 Êxito: {moeda(valor_exito)}")

    registrar_recibo(
        montar_recibo(
            titulo="RECIBO COMBINADO DE HONORÁRIOS",
            cliente=contrato["cliente"],
            documento=contrato["cpf_cnpj"],
            itens=itens,
            total=total,
            data=data_recebimento,
            metodo=forma,
            saldo_restante=novo_saldo,
        ),
        contrato["telefone"],
    )
    if novo_saldo is not None and novo_saldo <= 0:
        st.balloons()
    flash("Pagamento combinado registrado!")
    st.rerun()


# ----------------------------------------------------------- MEUS CONTRATOS --
def pagina_meus_contratos() -> None:
    st.header("Meus Contratos")
    st.markdown(
        "Acesse, edite e acompanhe os detalhes de cada contrato — "
        "incluindo parcelas da redução da liminar."
    )

    df_todos = select_db("SELECT * FROM contratos ORDER BY cliente ASC")
    if df_todos.empty:
        st.info("Nenhum contrato cadastrado ainda.")
        return

    numerico(df_todos, "valor_total", "saldo_devedor")

    busca = st.text_input("🔎 Buscar por nome, CPF/CNPJ ou nº do processo", placeholder="Digite para filtrar…")
    if busca.strip():
        termo = busca.strip().lower()
        termo_digitos = so_digitos(busca)
        def combina(linha: pd.Series) -> bool:
            alvos = [str(linha.get(campo, "")).lower() for campo in ("cliente", "nr_processo", "comarca")]
            if any(termo in alvo for alvo in alvos):
                return True
            return bool(termo_digitos) and termo_digitos in so_digitos(linha.get("cpf_cnpj", ""))

        df_todos = df_todos[df_todos.apply(combina, axis=1)]
        if df_todos.empty:
            st.warning("Nenhum contrato encontrado para esse filtro.")
            return

    mapa = _mapa_contratos(df_todos)
    rotulos = list(mapa)
    if st.session_state.get("meus_contratos_sel") not in rotulos:
        st.session_state.pop("meus_contratos_sel", None)

    rotulo = st.selectbox("Selecione o Contrato", options=rotulos, key="meus_contratos_sel")
    contrato_id = mapa[rotulo]
    contrato = df_todos[df_todos["id"] == contrato_id].iloc[0]

    caixa(
        f"<b>👤 {contrato['cliente']}</b> &nbsp;|&nbsp; "
        f"💳 {formatar_cpf_cnpj(contrato['cpf_cnpj'])} &nbsp;|&nbsp; "
        f"📞 {formatar_telefone(contrato['telefone'])} &nbsp;|&nbsp; "
        f"📅 Contrato: {formatar_data(contrato['data_contrato'])}"
    )
    processo = linha_processo(contrato)
    if processo:
        st.caption(processo)

    tutela = str(contrato.get("tutela", "") or "Pendente")
    df_resumo_liminar = select_db(
        "SELECT pago FROM parcelas_liminar WHERE contrato_id = %s", (contrato_id,)
    )
    if df_resumo_liminar.empty:
        rotulo_liminar, delta_liminar = "—", None
    else:
        numerico(df_resumo_liminar, "pago")
        pagas = int(df_resumo_liminar["pago"].sum())
        total = len(df_resumo_liminar)
        rotulo_liminar = f"{pagas} / {total}"
        delta_liminar = f"{total - pagas} pendente(s)" if pagas < total else "✅ Todas recebidas"

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Honorários Iniciais", moeda(contrato["valor_total"]))
    m2.metric("Saldo Devedor", moeda(contrato["saldo_devedor"]))
    m3.metric("Status da Tutela", tutela)
    m4.metric(
        "Parcelas da Redução",
        rotulo_liminar,
        delta=delta_liminar,
        delta_color="inverse" if (delta_liminar and "pendente" in delta_liminar) else "normal",
    )
    st.divider()

    _expander_editar(contrato_id, contrato, tutela)
    _expander_parcelas_iniciais(contrato_id, contrato)
    _expander_parcelas_liminar(contrato_id, contrato, tutela)


def _expander_editar(contrato_id: int, contrato: Any, tutela: str) -> None:
    with st.expander("✏️ Editar Contrato", expanded=False):
        st.subheader("Dados Gerais")
        col1, col2 = st.columns(2)
        nome = col1.text_input("Nome do Cliente", value=str(contrato["cliente"]), key=f"ed_nome_{contrato_id}")
        documento = col2.text_input("CPF / CNPJ", value=str(contrato["cpf_cnpj"] or ""), key=f"ed_cpf_{contrato_id}")
        telefone = col1.text_input("Telefone", value=str(contrato["telefone"] or ""), key=f"ed_tel_{contrato_id}")
        observacoes = col2.text_area(
            "Observações", value=str(contrato["observacoes"] or ""), key=f"ed_obs_{contrato_id}"
        )

        st.subheader("💰 Honorários Iniciais")
        col3, col4 = st.columns(2)
        ini_ativo = col3.selectbox(
            "Há cobrança inicial?", ["Não", "Sim"],
            index=0 if str(contrato.get("hon_inicial_ativo") or "Não") == "Não" else 1,
            key=f"ed_hi_ativo_{contrato_id}",
        )
        ini_valor = col4.number_input(
            "Valor Total dos Honorários Iniciais (R$)", min_value=0.0, step=100.0, format="%.2f",
            value=float(contrato.get("hon_inicial_valor") or 0), key=f"ed_hi_valor_{contrato_id}",
        )
        col5, col6 = st.columns(2)
        ini_parcelado = col5.selectbox(
            "Pagamento parcelado?", ["Não", "Sim"],
            index=0 if str(contrato.get("hon_inicial_parcelado") or "Não") == "Não" else 1,
            key=f"ed_hi_parc_{contrato_id}",
        )
        ini_qtd = int(
            col6.number_input(
                "Nº de Parcelas", min_value=1, max_value=60, step=1,
                value=int(contrato.get("hon_inicial_parcelas") or 1), key=f"ed_hi_qtd_{contrato_id}",
            )
        )
        col7, col8 = st.columns(2)
        valor_total = col7.number_input(
            "Valor Total do Contrato (R$)", min_value=0.0, step=100.0, format="%.2f",
            value=float(contrato["valor_total"]), key=f"ed_vt_{contrato_id}",
        )
        saldo_devedor = col8.number_input(
            "Saldo Devedor Atual (R$)", min_value=0.0, step=100.0, format="%.2f",
            value=float(contrato["saldo_devedor"]), key=f"ed_sd_{contrato_id}",
        )

        st.subheader("⚖️ Honorários da Liminar")
        col9, col10 = st.columns(2)
        tutela_nova = col9.selectbox(
            "Status da Tutela", STATUS_TUTELA,
            index=STATUS_TUTELA.index(tutela) if tutela in STATUS_TUTELA else 0,
            key=f"ed_tutela_{contrato_id}",
        )
        liminar_fixo = col10.number_input(
            "Honorários Fixos da Liminar (R$)", min_value=0.0, step=100.0, format="%.2f",
            value=float(contrato.get("hon_liminar_fixo") or 0), key=f"ed_lim_fixo_{contrato_id}",
        )
        col11, col12 = st.columns(2)
        reducao_valor = col11.number_input(
            "Valor da Redução Obtida (R$)", min_value=0.0, step=100.0, format="%.2f",
            value=float(contrato.get("hon_liminar_reducao_vlr") or 0), key=f"ed_red_vlr_{contrato_id}",
        )
        reducao_parcelas = int(
            col12.number_input(
                "Nº de Parcelas da Redução", min_value=0, max_value=360, step=1,
                value=int(contrato.get("hon_liminar_reducao_prc") or 0), key=f"ed_red_prc_{contrato_id}",
            )
        )

        st.subheader("🏆 Honorários de Êxito")
        col13, col14 = st.columns(2)
        exito_pct = col13.number_input(
            "Percentual de Êxito (%)", min_value=0.0, max_value=100.0, step=0.5, format="%.2f",
            value=float(contrato.get("hon_exito_percentual") or 0), key=f"ed_exito_pct_{contrato_id}",
        )
        exito_fixo = col14.number_input(
            "Valor Fixo de Êxito (R$)", min_value=0.0, step=100.0, format="%.2f",
            value=float(contrato.get("hon_exito_fixo") or 0), key=f"ed_exito_fixo_{contrato_id}",
        )

        st.subheader("📁 Dados do Processo")
        col15, col16 = st.columns(2)
        nr_processo = col15.text_input(
            "Número do Processo", value=str(contrato.get("nr_processo") or ""), key=f"ed_proc_{contrato_id}"
        )
        nr_vara = col16.text_input(
            "Número da Vara", value=str(contrato.get("nr_vara") or ""), key=f"ed_vara_{contrato_id}"
        )
        col17, col18 = st.columns(2)
        nome_juiz = col17.text_input(
            "Nome do Juiz", value=str(contrato.get("nome_juiz") or ""), key=f"ed_juiz_{contrato_id}"
        )
        comarca = col18.text_input(
            "Comarca", value=str(contrato.get("comarca") or ""), key=f"ed_com_{contrato_id}"
        )

        if not st.button("💾 Salvar Alterações", type="primary", key=f"ed_btn_{contrato_id}"):
            return

        if not nome.strip():
            st.error("Nome obrigatório.")
            return
        if so_digitos(documento):
            erro = validar_documento(documento)
            if erro:
                st.error(erro)
                return

        valor_parcela = round(ini_valor / ini_qtd, 2) if ini_qtd > 0 and ini_valor > 0 else 0.0
        exec_db(
            """
            UPDATE contratos SET
                cliente                 = %s,
                cpf_cnpj                = %s,
                telefone                = %s,
                observacoes             = %s,
                valor_total             = %s,
                saldo_devedor           = %s,
                hon_inicial_ativo       = %s,
                hon_inicial_valor       = %s,
                hon_inicial_parcelado   = %s,
                hon_inicial_parcelas    = %s,
                hon_inicial_vlr_parcela = %s,
                tutela                  = %s,
                hon_liminar_fixo        = %s,
                hon_liminar_reducao_vlr = %s,
                hon_liminar_reducao_prc = %s,
                hon_exito_percentual    = %s,
                hon_exito_fixo          = %s,
                nr_processo             = %s,
                nr_vara                 = %s,
                nome_juiz               = %s,
                comarca                 = %s,
                quitado_em              = CASE WHEN %s <= 0 THEN COALESCE(quitado_em, %s) END
            WHERE id = %s
            """,
            (
                nome.strip(),
                formatar_cpf_cnpj(so_digitos(documento)),
                telefone_para_banco(telefone),
                observacoes.strip() or None,
                valor_total,
                saldo_devedor,
                ini_ativo,
                ini_valor or None,
                ini_parcelado,
                ini_qtd or None,
                valor_parcela or None,
                tutela_nova,
                liminar_fixo or None,
                reducao_valor or None,
                reducao_parcelas or None,
                exito_pct or None,
                exito_fixo or None,
                nr_processo.strip() or None,
                nr_vara.strip() or None,
                nome_juiz.strip() or None,
                comarca.strip() or None,
                saldo_devedor,
                hoje().isoformat(),
                contrato_id,
            ),
        )
        flash("Contrato atualizado com sucesso!")
        st.rerun()


def _expander_parcelas_iniciais(contrato_id: int, contrato: Any) -> None:
    with st.expander("💰 Parcelas dos Honorários Iniciais", expanded=False):
        df = select_db(
            "SELECT * FROM parcelas WHERE contrato_id = %s ORDER BY nr_parcela", (contrato_id,)
        )

        if df.empty:
            st.info("Nenhuma parcela cadastrada para os honorários iniciais.")
        else:
            numerico(df, "pago", "valor_parcela")
            df["pago"] = df["pago"].astype(int)
            total = df["valor_parcela"].sum()
            pagas = df[df["pago"] == 1]["valor_parcela"].sum()

            col1, col2, col3 = st.columns(3)
            col1.metric("Total Parcelado", moeda(total))
            col2.metric("Já Recebido", moeda(pagas))
            col3.metric("A Receber", moeda(total - pagas))
            st.progress(
                float(pagas / total) if total > 0 else 0.0,
                text=f"Progresso: {porcentagem(pagas / total if total else 0)} recebido",
            )
            _tabela_parcelas(df, "data_vencimento", "Vencimento")

            # Estorno: antes só era possível corrigir um erro de baixa
            # editando o banco na mão.
            baixadas = df[df["pago"] == 1]
            if not baixadas.empty:
                st.markdown("**Estornar uma baixa feita por engano:**")
                opcoes = {
                    f"Parcela {linha.nr_parcela} — {moeda(linha.valor_parcela)} "
                    f"(pago em {formatar_data(linha.data_pagamento)})": int(linha.nr_parcela)
                    for linha in baixadas.itertuples()
                }
                col_a, col_b = st.columns([3, 1])
                numero = opcoes[col_a.selectbox("Parcela:", list(opcoes), key=f"est_sel_{contrato_id}")]
                confirmar = col_b.checkbox("Confirmo", key=f"est_conf_{contrato_id}")
                if st.button("↩️ Estornar Parcela", key=f"est_btn_{contrato_id}", disabled=not confirmar):
                    _estornar_parcela_inicial(contrato_id, numero)
                    flash(f"Parcela {numero} estornada. O saldo devedor foi recomposto.", "warning")
                    st.rerun()

        st.divider()
        st.markdown("**Recriar o parcelamento** (apaga as parcelas atuais e gera novas):")
        valor_base = float(contrato["valor_total"]) or 0.0
        col1, col2, col3 = st.columns(3)
        novo_total = col1.number_input(
            "Valor a parcelar (R$)", min_value=0.01, step=100.0, format="%.2f",
            value=valor_base if valor_base > 0 else 100.0, key=f"rec_vlr_{contrato_id}",
        )
        nova_qtd = int(
            col2.number_input(
                "Nº de parcelas", min_value=1, max_value=60, step=1,
                value=int(contrato.get("hon_inicial_parcelas") or 1), key=f"rec_qtd_{contrato_id}",
            )
        )
        primeiro = col3.date_input(
            "Vencimento da 1ª", value=hoje(), key=f"rec_venc_{contrato_id}", format=FORMATO_DATA_WIDGET)
        st.caption(f"Cada parcela ficará em **{moeda(novo_total / max(nova_qtd, 1))}**.")

        confirmar_recriar = st.checkbox(
            "Entendo que as parcelas atuais (e suas baixas) serão apagadas.",
            key=f"rec_conf_{contrato_id}",
        )
        if st.button(
            "🔁 Recriar Parcelas", key=f"rec_btn_{contrato_id}", disabled=not confirmar_recriar
        ):
            _recriar_parcelas(contrato_id, novo_total, nova_qtd, primeiro)
            flash(f"{nova_qtd} parcela(s) recriada(s). Saldo devedor reajustado.")
            st.rerun()


def _estornar_parcela_inicial(contrato_id: int, numero: int) -> None:
    with transacao() as cur:
        cur.execute(
            "SELECT valor_parcela FROM parcelas WHERE contrato_id = %s AND nr_parcela = %s AND pago = 1",
            (contrato_id, numero),
        )
        linha = cur.fetchone()
        if not linha:
            raise ErroBanco("Parcela não encontrada ou já está pendente.")
        valor = float(linha[0] or 0)

        cur.execute(
            """UPDATE parcelas
                  SET pago = 0, data_pagamento = NULL, forma_pagamento = NULL
                WHERE contrato_id = %s AND nr_parcela = %s""",
            (contrato_id, numero),
        )
        cur.execute(
            """UPDATE contratos
                  SET saldo_devedor = LEAST(COALESCE(saldo_devedor, 0) + %s, COALESCE(valor_total, 0)),
                      quitado_em = NULL
                WHERE id = %s""",
            (valor, contrato_id),
        )


def _recriar_parcelas(contrato_id: int, total: float, quantidade: int, primeiro: date) -> None:
    valores = dividir_parcelas(total, quantidade)
    vencimentos = gerar_vencimentos(primeiro, quantidade)
    with transacao() as cur:
        cur.execute("DELETE FROM parcelas WHERE contrato_id = %s", (contrato_id,))
        execute_values(
            cur,
            "INSERT INTO parcelas (contrato_id, nr_parcela, valor_parcela, data_vencimento) VALUES %s",
            [
                (contrato_id, numero, valor, vencimento.strftime("%Y-%m-%d"))
                for numero, (valor, vencimento) in enumerate(zip(valores, vencimentos), start=1)
            ],
        )
        cur.execute(
            """UPDATE contratos
                  SET valor_total = %s, saldo_devedor = %s,
                      hon_inicial_parcelas = %s, hon_inicial_vlr_parcela = %s,
                      hon_inicial_parcelado = CASE WHEN %s > 1 THEN 'Sim' ELSE 'Não' END,
                      quitado_em = NULL
                WHERE id = %s""",
            (total, total, quantidade, round(total / quantidade, 2), quantidade, contrato_id),
        )


def _expander_parcelas_liminar(contrato_id: int, contrato: Any, tutela: str) -> None:
    with st.expander("📋 Parcelas da Redução da Liminar", expanded=True):
        df = select_db(
            "SELECT * FROM parcelas_liminar WHERE contrato_id = %s ORDER BY nr_parcela", (contrato_id,)
        )

        if df.empty:
            st.info("Nenhuma parcela da redução cadastrada para este contrato.")
            if tutela not in TUTELA_COM_REDUCAO:
                st.warning(
                    "A tutela ainda está como **Pendente** ou **Indeferida**. "
                    "Edite o status da tutela acima para cadastrar parcelas da redução."
                )
                return

            st.markdown("**Cadastrar parcelas da redução obtida:**")
            col1, col2, col3 = st.columns(3)
            total = col1.number_input(
                "Valor Total da Redução (R$)", min_value=0.01, step=100.0, format="%.2f",
                value=float(contrato.get("hon_liminar_reducao_vlr") or 0) or 100.0,
                key=f"pl_total_{contrato_id}",
            )
            quantidade = int(
                col2.number_input(
                    "Número de Parcelas", min_value=1, max_value=360, step=1,
                    value=max(int(contrato.get("hon_liminar_reducao_prc") or 1), 1),
                    key=f"pl_qtd_{contrato_id}",
                )
            )
            inicio = col3.date_input("Data da 1ª Parcela", value=hoje(), key=f"pl_inicio_{contrato_id}", format=FORMATO_DATA_WIDGET)
            resumo_parcelamento(total, quantidade, "Total da redução")

            if st.button("📥 Criar Parcelas da Redução", type="primary", key=f"pl_btn_{contrato_id}"):
                valores = dividir_parcelas(total, quantidade)
                vencimentos = gerar_vencimentos(inicio, quantidade)
                with transacao() as cur:
                    execute_values(
                        cur,
                        """INSERT INTO parcelas_liminar
                           (contrato_id, nr_parcela, valor_parcela, data_prevista) VALUES %s""",
                        [
                            (contrato_id, numero, valor, vencimento.strftime("%Y-%m-%d"))
                            for numero, (valor, vencimento) in enumerate(zip(valores, vencimentos), start=1)
                        ],
                    )
                flash(f"{quantidade} parcela(s) criada(s) com sucesso!")
                st.rerun()
            return

        numerico(df, "pago", "valor_parcela")
        df["pago"] = df["pago"].astype(int)
        total = df["valor_parcela"].sum()
        recebido = df[df["pago"] == 1]["valor_parcela"].sum()

        col1, col2, col3 = st.columns(3)
        col1.metric("Total da Redução", moeda(total))
        col2.metric("Já Recebido", moeda(recebido))
        col3.metric("A Receber", moeda(total - recebido))
        proporcao = float(recebido / total) if total > 0 else 0.0
        st.progress(proporcao, text=f"Progresso: {porcentagem(proporcao)} recebido")

        _tabela_parcelas(df, "data_prevista", "Previsão")

        pendentes = df[df["pago"] == 0]
        if pendentes.empty:
            st.success("🎉 Todas as parcelas da redução já foram recebidas!")
        else:
            st.markdown("**Registrar Recebimento:**")
            col_a, col_b = st.columns(2)
            opcoes = {
                f"Parcela {linha.nr_parcela} — Prev. {formatar_data(linha.data_prevista)} — "
                f"{moeda(linha.valor_parcela)}": int(linha.nr_parcela)
                for linha in pendentes.itertuples()
            }
            numero = opcoes[col_a.selectbox("Qual parcela recebeu?", list(opcoes), key=f"mc_lim_sel_{contrato_id}")]
            data_recebimento = col_b.date_input(
                "Data do Recebimento", value=hoje(), key=f"mc_lim_data_{contrato_id}", format=FORMATO_DATA_WIDGET)
            # Aqui não há campo de valor: a baixa registra a parcela prevista.
            # Para receber valor diferente do previsto e emitir recibo, use
            # 💰 Pagamentos → aba Liminar / Redução.
            st.caption(
                f"Valor previsto da parcela {numero}: "
                f"**{moeda(pendentes[pendentes['nr_parcela'] == numero]['valor_parcela'].iloc[0])}**"
            )
            if st.button("✅ Confirmar Recebimento da Parcela", type="primary", key=f"mc_lim_btn_{contrato_id}"):
                _baixar_parcela_liminar(contrato_id, numero, data_recebimento)
                flash(f"Parcela {numero} marcada como recebida!")
                st.rerun()

        st.divider()
        confirmar = st.checkbox(
            "Confirmo que quero apagar TODAS as parcelas da redução deste contrato.",
            key=f"pl_del_conf_{contrato_id}",
        )
        if st.button(
            "🗑️ Apagar todas as parcelas da redução", key=f"pl_del_{contrato_id}", disabled=not confirmar
        ):
            exec_db("DELETE FROM parcelas_liminar WHERE contrato_id = %s", (contrato_id,))
            flash("Parcelas da redução removidas.", "warning")
            st.rerun()


# ---------------------------------------------------------------- ARQUIVADOS --
def pagina_arquivados() -> None:
    st.header("Contratos Quitados")
    st.markdown("Histórico de clientes que já **zeraram** seus saldos devedores.")

    df = select_db(
        """
        SELECT c.cliente, c.cpf_cnpj, c.telefone, c.valor_total, c.data_contrato,
               COALESCE(c.quitado_em::text, MAX(p.data_pagamento)::text) AS data_quitacao,
               c.observacoes
        FROM contratos c
        LEFT JOIN parcelas p ON c.id = p.contrato_id
        WHERE c.saldo_devedor <= 0
        GROUP BY c.id
        ORDER BY c.cliente ASC
        """
    )
    if df.empty:
        st.info("Nenhum contrato arquivado até o momento.")
        return

    numerico(df, "valor_total")
    visao = pd.DataFrame(
        {
            "Cliente": df["cliente"],
            "CPF/CNPJ": df["cpf_cnpj"].apply(formatar_cpf_cnpj),
            "Telefone": df["telefone"].apply(formatar_telefone),
            "Valor Total": df["valor_total"],
            "Data Início": df["data_contrato"].apply(formatar_data),
            "Data Quitação": df["data_quitacao"].apply(formatar_data),
            "Observações": df["observacoes"].apply(lambda v: "-" if nulo(v) else str(v)),
        }
    )
    tabela(visao, ["Valor Total"])
    st.divider()
    botoes_exportacao(visao, "contratos_quitados", "Relatório de Contratos Quitados")


# -------------------------------------------------------------------- GESTÃO --
def pagina_gestao() -> None:
    st.header("Gerenciar")

    aba_excluir, aba_backup, aba_diagnostico = st.tabs(
        ["🗑️ Excluir Contrato", "💾 Backup", "🔧 Diagnóstico"]
    )

    with aba_excluir:
        df = select_db("SELECT id, cliente, saldo_devedor FROM contratos ORDER BY cliente ASC")
        if df.empty:
            st.info("Nenhum contrato cadastrado.")
        else:
            opcoes = {f"{linha['cliente']} (ID {linha['id']})": int(linha["id"]) for _, linha in df.iterrows()}
            rotulo = st.selectbox("Excluir contrato:", options=list(opcoes))
            contrato_id = opcoes[rotulo]
            nome_esperado = str(df[df["id"] == contrato_id].iloc[0]["cliente"]).strip()

            st.warning(
                "A exclusão remove o contrato **e todas as parcelas vinculadas** "
                "(iniciais e da liminar). Não há como desfazer."
            )
            digitado = st.text_input(
                f"Para confirmar, digite o nome do cliente exatamente: **{nome_esperado}**",
                key="gestao_confirma_nome",
            )
            liberado = digitado.strip() == nome_esperado
            if st.button("❌ APAGAR DEFINITIVAMENTE", disabled=not liberado):
                exec_db("DELETE FROM contratos WHERE id = %s", (contrato_id,))
                flash(f"Contrato de {nome_esperado} excluído.", "warning")
                st.rerun()

    with aba_backup:
        st.markdown(
            "Baixe uma cópia completa das três tabelas em um único arquivo Excel "
            "(uma aba por tabela). Guarde periodicamente fora do Supabase."
        )
        if st.button("Gerar backup agora"):
            st.session_state["gerar_backup"] = True
        if st.session_state.get("gerar_backup"):
            dados = {
                "contratos": select_db("SELECT * FROM contratos ORDER BY id", cache=False),
                "parcelas": select_db("SELECT * FROM parcelas ORDER BY contrato_id, nr_parcela", cache=False),
                "parcelas_liminar": select_db(
                    "SELECT * FROM parcelas_liminar ORDER BY contrato_id, nr_parcela", cache=False
                ),
            }
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                # `conteudo`, nao `tabela`: usar o nome da funcao `tabela()` como
                # variavel de laco faz o Python tratar o nome como local em toda a
                # funcao, quebrando a chamada real la na aba de Diagnostico.
                for nome, conteudo in dados.items():
                    (conteudo if not conteudo.empty else pd.DataFrame({"vazio": []})).to_excel(
                        writer, index=False, sheet_name=nome[:31]
                    )
            st.download_button(
                "📥 Baixar backup completo",
                data=buffer.getvalue(),
                file_name=f"backup_honorarios_{hoje()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
            st.caption(
                " • ".join(f"{nome}: {len(conteudo)} registro(s)" for nome, conteudo in dados.items())
            )

    with aba_diagnostico:
        try:
            for aviso in inicializar_banco()["avisos"]:
                st.warning(aviso)
        except ErroBanco as erro:
            st.error(f"Estrutura do banco indisponível: {str(erro).strip()[:300]}")

        col1, col2, col3 = st.columns(3)
        col1.metric("Contratos", int(escalar("SELECT COUNT(*) FROM contratos")))
        col2.metric("Parcelas (iniciais)", int(escalar("SELECT COUNT(*) FROM parcelas")))
        col3.metric("Parcelas (liminar)", int(escalar("SELECT COUNT(*) FROM parcelas_liminar")))

        parametros = _parametros_conexao()
        st.write(
            {
                "host": parametros["host"],
                "porta": parametros["port"],
                "banco": parametros["dbname"],
                "fuso": str(FUSO),
                "pdf_disponivel": PDF_DISPONIVEL,
                "keepalive_ativo": SCHEDULER_DISPONIVEL,
                "cache_ttl_segundos": CACHE_TTL,
                "versao_cache": _estado_cache()["versao"],
            }
        )
        st.caption(f"Servidor: {escalar('SELECT version()', padrao='—')}")

        # O tipo real das colunas nem sempre bate com o que o CREATE TABLE do
        # código declara: o DDL só vale para tabela nova, então uma coluna
        # alterada à mão no Supabase fica invisível até quebrar uma consulta.
        with st.expander("🧬 Estrutura real das tabelas", expanded=False):
            st.caption(
                "Tipos como estão no banco agora. Útil quando uma consulta falha "
                "com erro de tipo — foi assim que apareceu o `timestamp` onde o "
                "código esperava `text`."
            )
            try:
                estrutura = select_db(SQL_ESTRUTURA, cache=False)
                if estrutura.empty:
                    st.info("Nenhuma das três tabelas foi encontrada no schema `public`.")
                else:
                    estrutura.columns = ["Tabela", "Coluna", "Tipo", "Aceita nulo", "Padrão"]
                    tabela(estrutura, altura=420)
            except ErroBanco as erro:
                st.error(f"Não foi possível ler a estrutura: {erro}")

        if st.button("♻️ Limpar todo o cache de consultas"):
            st.cache_data.clear()
            _invalidar_cache()
            flash("Cache limpo.", "info")
            st.rerun()


# =============================================================================
# 12. ROTEAMENTO
# =============================================================================
PAGINAS: dict[str, Callable[[], None]] = {
    "📊 Dashboard": pagina_dashboard,
    "➕ Novo Contrato": pagina_novo_contrato,
    "💰 Pagamentos": pagina_pagamentos,
    "📂 Meus Contratos": pagina_meus_contratos,
    "📁 Arquivados": pagina_arquivados,
    "⚙️ Gestão": pagina_gestao,
}


def barra_lateral() -> str:
    with st.sidebar:
        simbolo = _logo_embutido("symbol-color.png")
        if simbolo:
            st.markdown(
                f"""
                <div style='display:flex;align-items:center;gap:11px;padding:6px 0 14px;'>
                    <img src='data:image/png;base64,{simbolo}' alt='' style='height:34px;'>
                    <div style='font-family:Fraunces,Georgia,serif;line-height:1.25;'>
                        <div style='font-size:.92rem;color:#012B34;'>Maciel Freitas</div>
                        <div style='font-size:.7rem;color:#8B9296;letter-spacing:.08em;'>ADVOCACIA</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
        else:
            st.markdown(f"### {APP_ICONE} Honorários")

        usuario = st.session_state.get("usuario")
        if usuario:
            st.caption(f"Conectado como **{usuario}**")
        st.divider()

        aba = st.radio("Navegação", MENU, key="rad_nav")

        st.divider()
        if st.button("🔄 Atualizar dados", use_container_width=True):
            _invalidar_cache()
            st.rerun()
        if st.button("🚪 Sair", use_container_width=True):
            st.session_state.clear()
            st.rerun()
        st.caption(f"🕒 {agora().strftime(FORMATO_DATA)} {agora():%H:%M}")
    return aba


def main() -> None:
    # Antes do login de propósito: qualquer carregamento da página, mesmo sem
    # ninguém entrar, já mantém o Supabase marcado como ativo. Se isso ficasse
    # depois do `autenticar()`, uma semana sem login bastaria para o projeto
    # ser pausado por inatividade.
    iniciar_keepalive()

    if not autenticar():
        st.stop()

    try:
        estado = inicializar_banco()
    except ConfiguracaoAusente as erro:
        st.error(f"⚙️ {erro}")
        st.stop()
    except ErroBanco as erro:
        st.error("❌ Não foi possível conectar ao banco de dados.")
        st.caption(str(erro).strip()[:400])
        if st.button("🔄 Tentar novamente", type="primary"):
            st.rerun()
        st.stop()

    for aviso in estado["avisos"]:
        st.sidebar.warning(aviso, icon="⚠️")

    if "rad_nav" not in st.session_state:
        st.session_state["rad_nav"] = MENU[0]

    aba = barra_lateral()
    cabecalho_marca(aba.split(" ", 1)[-1])
    mostrar_flash()

    try:
        PAGINAS[aba]()
    except ConfiguracaoAusente as erro:
        st.error(f"⚙️ {erro}")
    except ErroBanco as erro:
        st.error(f"❌ Erro no banco de dados: {erro}")
        st.caption("Tente novamente. Se persistir, verifique o Supabase em ⚙️ Gestão → Diagnóstico.")
    except Exception as erro:  # noqa: BLE001 - rede de segurança da UI
        # Resumo controlado: aparece mesmo com showErrorDetails=false no
        # config.toml. O traceback completo fica atrás do expander e nos logs.
        st.error("❌ Ocorreu um erro inesperado nesta tela.")
        st.caption(f"{type(erro).__name__}: {str(erro).strip()[:300]}")
        with st.expander("Detalhes técnicos"):
            st.exception(erro)


main()
